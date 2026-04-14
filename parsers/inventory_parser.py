"""Parser for inventory (інвентаризація) Excel files."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from models import InventoryItem, ColumnMapping


def col_letter_to_index(letter: str) -> int:
    """Convert column letter (A, B, ..., Z, AA, ...) to 1-based index."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def parse_inventory(
    filepath: str | Path,
    mapping: ColumnMapping | None = None,
    sheet_name: str | None = None,
) -> list[InventoryItem]:
    """Parse an inventory Excel file into structured InventoryItem objects.

    Args:
        filepath: Path to the Excel file.
        mapping: Column mapping. If None, uses Chirana defaults.
        sheet_name: Specific sheet to parse. If None, uses the first sheet.

    Returns:
        List of InventoryItem objects.
    """
    if mapping is None:
        mapping = ColumnMapping(
            number="B",
            article="D",
            name="F",
            deviation="K",
            quantity="N",           # fact quantity
            quantity_accounting="Q",
            unit="T",
            price="V",
            header_row=11,
            data_start_row=12,
        )

    low_path = str(filepath).lower()
    looks_fakturace = "fakturace" in low_path or "soupis" in low_path
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=not looks_fakturace)
    try:
        primary_sheet_name = sheet_name or _select_primary_inventory_sheet(wb, mapping)
        if mapping is not None:
            mapping.sheet_name = primary_sheet_name
        ws = wb[primary_sheet_name]
        adaptive_mapping = _detect_inventory_mapping_from_headers(ws)
        active_mapping = mapping
        if adaptive_mapping and not _mapping_has_enough_rows(ws, mapping):
            active_mapping = adaptive_mapping
            active_mapping.sheet_name = primary_sheet_name
            mapping.sheet_name = primary_sheet_name
        items = _parse_inventory_sheet_with_mapping(ws, active_mapping)
        # Fallback for Fakturace-like sheets where classic mapping yields nothing useful.
        if len(items) < 10 or looks_fakturace or "soupis prac" in _sheet_signature(ws):
            fallback_items = _parse_fakturace_like_workbook(wb)
            if len(fallback_items) > len(items):
                return fallback_items
        return items
    finally:
        wb.close()


def get_inventory_preview(
    filepath: str | Path,
    num_rows: int = 15,
    sheet_name: str | None = None,
) -> list[list[Any]]:
    """Get first N rows from the inventory file for preview/mapping."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(max_row=num_rows)):
        rows.append([cell.value for cell in row])

    wb.close()
    return rows


def _sheet_signature(ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
    parts: list[str] = []
    for row_vals in ws.iter_rows(min_row=1, max_row=15, max_col=min(ws.max_column, 30), values_only=True):
        text = " ".join(str(v) for v in row_vals if v not in (None, ""))
        if text:
            parts.append(text.lower())
    return " | ".join(parts)


def _select_primary_inventory_sheet(
    wb: openpyxl.Workbook,
    mapping: ColumnMapping,
) -> str:
    scored: list[tuple[int, str]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        score = 0
        sig = _sheet_signature(ws)
        if any(k in sig for k in ["invent", "sklad", "zásob", "zasob"]):
            score += 40
        if any(k in sig for k in ["pč", "typ", "kód", "kod", "popis", "množství", "mnozstvi", "j.cena"]):
            score += 35
        if any(k in name.lower() for k in ["rekap", "summary"]):
            score -= 20
        # Check if mapped name column has values in expected data rows.
        idx = col_letter_to_index(mapping.name or "F")
        non_empty = 0
        for row in ws.iter_rows(min_row=max(1, mapping.data_start_row), max_row=min(ws.max_row, mapping.data_start_row + 40)):
            if idx - 1 < len(row):
                val = row[idx - 1].value
                if val not in (None, ""):
                    non_empty += 1
        score += min(non_empty, 30)
        scored.append((score, name))
    scored.sort(reverse=True)
    return scored[0][1] if scored else wb.active.title


def _detect_inventory_mapping_from_headers(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> ColumnMapping | None:
    max_col = min(ws.max_column, 40)
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), max_col=max_col, values_only=True))

    def normalize(val: Any) -> str:
        return str(val or "").strip().lower()

    for row_idx, row_vals in enumerate(rows, 1):
        next_vals = rows[row_idx] if row_idx < len(rows) else ()
        row_text = " | ".join(normalize(v) for v in row_vals if v not in (None, ""))
        if not any(key in row_text for key in ("товар", "název", "nazev", "popis", "матеріал", "material")):
            continue

        def find_col(keys: tuple[str, ...]) -> int | None:
            for col_idx in range(1, max_col + 1):
                top = normalize(row_vals[col_idx - 1] if col_idx - 1 < len(row_vals) else None)
                bottom = normalize(next_vals[col_idx - 1] if col_idx - 1 < len(next_vals) else None)
                combined = f"{top} {bottom}".strip()
                if any(key in combined for key in keys):
                    return col_idx
            return None

        name_col = find_col(("товар", "název", "nazev", "popis", "матеріал", "material"))
        if not name_col:
            continue

        row_col = find_col(("№", "č.", "pč", "por."))
        deviation_col = find_col(("відхил", "odchyl", "deviation", "rozdíl", "rozdil"))
        quantity_col = find_col(("кількість", "množství", "mnozstvi", "fact", "факт"))
        unit_col = find_col(("mj", "jed", "unit", "од."))
        article_col = find_col(("код", "kód", "kod", "article", "artikl"))
        price_col = find_col(("цена", "cena", "price", "j.cena", "jedn.cena"))

        data_start_row = row_idx + 2 if any(normalize(v) for v in next_vals) else row_idx + 1
        return ColumnMapping(
            row_number=openpyxl.utils.get_column_letter(row_col) if row_col else None,
            article=openpyxl.utils.get_column_letter(article_col) if article_col else None,
            name=openpyxl.utils.get_column_letter(name_col),
            unit=openpyxl.utils.get_column_letter(unit_col) if unit_col else None,
            quantity=openpyxl.utils.get_column_letter(quantity_col) if quantity_col else None,
            quantity_accounting=None,
            deviation=openpyxl.utils.get_column_letter(deviation_col) if deviation_col else None,
            price=openpyxl.utils.get_column_letter(price_col) if price_col else None,
            total=None,
            header_row=row_idx,
            data_start_row=data_start_row,
        )
    return None


def _mapping_has_enough_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    mapping: ColumnMapping,
    min_rows: int = 5,
) -> bool:
    if not mapping.name:
        return False
    idx = col_letter_to_index(mapping.name) - 1
    non_empty = 0
    for row in ws.iter_rows(
        min_row=max(1, mapping.data_start_row),
        max_row=min(ws.max_row, mapping.data_start_row + 60),
        values_only=True,
    ):
        if idx < len(row):
            val = row[idx]
            if val not in (None, "") and str(val).strip():
                non_empty += 1
                if non_empty >= min_rows:
                    return True
    return False


def _to_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        cleaned = str(val).replace(",", ".").replace(" ", "").replace("\xa0", "").strip()
        if cleaned in ("", "-", "—"):
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _parse_inventory_sheet_with_mapping(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    mapping: ColumnMapping,
) -> list[InventoryItem]:
    items: list[InventoryItem] = []

    def get_cell(row_data: tuple, col_letter: str | None) -> Any:
        if not col_letter:
            return None
        idx = col_letter_to_index(col_letter) - 1
        if idx < len(row_data):
            cell = row_data[idx]
            try:
                return cell.value if cell else None
            except AttributeError:
                return None
        return None

    row_num = mapping.data_start_row - 1
    for row in ws.iter_rows(min_row=mapping.data_start_row):
        row_num += 1
        try:
            if row[0] and hasattr(row[0], "row") and row[0].row:
                row_num = row[0].row
        except (AttributeError, TypeError):
            pass

        name_val = get_cell(row, mapping.name)
        if not name_val or not str(name_val).strip():
            continue

        name_str = str(name_val).strip()
        article_val = get_cell(row, mapping.article)
        article_str = str(article_val).strip() if article_val else None

        deviation = _to_float(get_cell(row, mapping.deviation))
        quantity_fact = _to_float(get_cell(row, mapping.quantity))
        # Fallback: many formats do not have explicit "deviation", use quantity as factual amount.
        if deviation is None and quantity_fact is not None:
            deviation = quantity_fact

        item = InventoryItem(
            row=row_num,
            number=str(get_cell(row, mapping.row_number) or "").strip() or None,
            article=article_str if article_str else None,
            name=name_str,
            unit=str(get_cell(row, mapping.unit) or "").strip() or None,
            quantity_fact=quantity_fact,
            quantity_accounting=_to_float(get_cell(row, mapping.quantity_accounting)),
            deviation=deviation,
            price=_to_float(get_cell(row, mapping.price)),
        )
        items.append(item)
    return items


def _parse_fakturace_like_workbook(wb: openpyxl.Workbook) -> list[InventoryItem]:
    all_items: list[InventoryItem] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_map = _find_fakturace_header(ws)
        if not header_map:
            continue
        all_items.extend(_parse_fakturace_sheet(ws, header_map))
    return all_items


def _find_fakturace_header(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int] | None:
    max_col = min(ws.max_column, 80)
    for row_idx in range(1, min(ws.max_row, 220) + 1):
        row_vals = [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]
        norm = [str(v).strip().lower() if v not in (None, "") else "" for v in row_vals]
        if not norm:
            continue
        def find_col(keys: tuple[str, ...]) -> int | None:
            for i, val in enumerate(norm, 1):
                if any(k in val for k in keys):
                    return i
            return None
        col_typ = find_col(("typ",))
        col_kod = find_col(("kód", "kod"))
        col_popis = find_col(("popis", "název", "nazev"))
        col_mj = find_col(("mj", "jed", "jedn"))
        col_qty = find_col(("množství", "mnozstvi"))
        col_price = find_col(("j.cena", "jedn.cena", "jednotk"))
        col_total = find_col(("cena celkem",))
        if col_typ and col_kod and col_popis and col_qty:
            return {
                "header_row": row_idx,
                "typ": col_typ,
                "kod": col_kod,
                "popis": col_popis,
                "mj": col_mj or 0,
                "qty": col_qty,
                "price": col_price or 0,
                "total": col_total or 0,
            }
    return None


def _parse_fakturace_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_map: dict[str, int],
) -> list[InventoryItem]:
    result: list[InventoryItem] = []
    blank_streak = 0
    for row_idx in range(header_map["header_row"] + 1, ws.max_row + 1):
        typ_val = ws.cell(row_idx, header_map["typ"]).value
        code_val = ws.cell(row_idx, header_map["kod"]).value
        name_val = ws.cell(row_idx, header_map["popis"]).value
        if name_val in (None, "") and code_val in (None, ""):
            blank_streak += 1
            if blank_streak >= 60:
                break
            continue
        blank_streak = 0
        typ = str(typ_val or "").strip().upper()
        # Keep actual line items, ignore aggregates/expressions.
        if typ in {"", "D", "VV"}:
            continue
        name = str(name_val or "").strip()
        if not name:
            continue
        qty = _to_float(ws.cell(row_idx, header_map["qty"]).value)
        price = _to_float(ws.cell(row_idx, header_map["price"]).value) if header_map["price"] else None
        total = _to_float(ws.cell(row_idx, header_map["total"]).value) if header_map["total"] else None
        deviation = qty if qty is not None else total
        result.append(
            InventoryItem(
                row=row_idx,
                number=None,
                article=str(code_val).strip() if code_val not in (None, "") else None,
                name=name,
                unit=str(ws.cell(row_idx, header_map["mj"]).value).strip() if header_map["mj"] and ws.cell(row_idx, header_map["mj"]).value not in (None, "") else None,
                quantity_fact=qty,
                quantity_accounting=None,
                deviation=deviation,
                price=price,
            )
        )
    return result
