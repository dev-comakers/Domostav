"""AI-powered column mapping engine for unknown Excel structures."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from rich.console import Console
from rich.table import Table

from models import ColumnMapping
from llm.client import ClaudeClient

console = Console()

MAPPING_PROMPT = """Analyze these first rows of an Excel file and determine the column mapping.

The file contains construction material data. I need you to identify which columns contain:
- row_number: Sequential row number
- article: Article/material code (like STRE020S4)
- name: Material or work item name/description
- unit: Unit of measurement (m, ks, bm, kg, etc.)
- quantity: Quantity (actual/fact)
- quantity_accounting: Quantity per accounting records
- deviation: Deviation (difference between fact and accounting)
- price: Price per unit
- total: Total amount
- percent_month: Percentage completed this month
- total_month: Total amount for this month

Also identify:
- header_row: Which row number contains the column headers
- data_start_row: Which row number is the first data row

Here are the first rows of the file:
{preview_json}

Respond with a JSON object with column letters (A, B, C, ...) for each field, or null if not found.
Example: {{"name": "F", "article": "D", "unit": "K", "quantity": "N", "header_row": 5, "data_start_row": 6}}
"""


def get_excel_preview(
    filepath: str | Path,
    num_rows: int = 18,
    file_type: str = "inventory",
) -> list[dict]:
    """Get first rows from the most data-rich sheet for mapping."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    ws = _pick_best_sheet_for_mapping(wb, file_type=file_type)
    rows = []
    for i, row in enumerate(ws.iter_rows(max_row=num_rows), 1):
        cols = {}
        for j, cell in enumerate(row):
            col_letter = chr(ord("A") + j) if j < 26 else chr(ord("A") + j // 26 - 1) + chr(ord("A") + j % 26)
            if cell.value is not None:
                cols[col_letter] = str(cell.value)[:100]
        if cols:
            rows.append({"row": i, "columns": cols})
    wb.close()
    return rows


def _pick_best_sheet_for_mapping(
    wb: openpyxl.Workbook,
    file_type: str = "inventory",
) -> openpyxl.worksheet.worksheet.Worksheet:
    scored: list[tuple[int, str]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        score = 0
        if "rekap" in name.lower() or "summary" in name.lower():
            score -= 30
        max_col = min(ws.max_column, 60)
        non_empty_cells = 0
        keyword_hits = 0
        for row_idx in range(1, min(ws.max_row, 60) + 1):
            row_values = [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]
            text = " | ".join(str(v) for v in row_values if v not in (None, "")).lower()
            non_empty_cells += sum(1 for v in row_values if v not in (None, ""))
            if file_type == "spp":
                if any(k in text for k in ["název položky", "nazev polozky", "množství", "mnozstvi", "montáž", "celkem"]):
                    keyword_hits += 1
            else:
                if any(k in text for k in ["article", "artikl", "název", "nazev", "popis", "množství", "mnozstvi", "deviation", "odchyl"]):
                    keyword_hits += 1
        score += min(non_empty_cells // 10, 100)
        score += keyword_hits * 15
        scored.append((score, name))
    scored.sort(reverse=True)
    return wb[scored[0][1]] if scored else wb.active


def auto_detect_mapping(
    filepath: str | Path,
    client: ClaudeClient,
    file_type: str = "inventory",
) -> ColumnMapping:
    """Use Claude to detect column mapping for an unknown Excel file.

    Args:
        filepath: Path to the Excel file.
        client: ClaudeClient instance.
        file_type: "inventory" or "spp" — helps the AI understand context.

    Returns:
        Detected ColumnMapping.
    """
    preview = get_excel_preview(filepath, file_type=file_type)
    preview_json = json.dumps(preview, ensure_ascii=False, indent=2)

    prompt = MAPPING_PROMPT.format(preview_json=preview_json)
    if file_type == "spp":
        prompt += "\nThis is a file of performed construction works (SPP), not inventory."
    else:
        prompt += "\nThis is an inventory/stock-taking file."

    result = client.ask_json(prompt)

    return ColumnMapping(
        row_number=result.get("row_number"),
        article=result.get("article"),
        name=result.get("name"),
        unit=result.get("unit"),
        quantity=result.get("quantity"),
        quantity_accounting=result.get("quantity_accounting"),
        deviation=result.get("deviation"),
        price=result.get("price"),
        total=result.get("total"),
        percent_month=result.get("percent_month"),
        total_month=result.get("total_month"),
        header_row=result.get("header_row", 1),
        data_start_row=result.get("data_start_row", 2),
    )


def display_mapping(mapping: ColumnMapping, filepath: str) -> None:
    """Display detected mapping in a nice table for user confirmation."""
    table = Table(title=f"Detected Column Mapping: {Path(filepath).name}")
    table.add_column("Field", style="cyan")
    table.add_column("Column", style="green")

    fields = [
        ("Name", mapping.name),
        ("Article", mapping.article),
        ("Unit", mapping.unit),
        ("Quantity (fact)", mapping.quantity),
        ("Quantity (accounting)", mapping.quantity_accounting),
        ("Deviation", mapping.deviation),
        ("Price", mapping.price),
        ("Total", mapping.total),
        ("% Month", mapping.percent_month),
        ("Total Month", mapping.total_month),
        ("Header Row", str(mapping.header_row)),
        ("Data Start Row", str(mapping.data_start_row)),
    ]
    for name, val in fields:
        table.add_row(name, val or "—")

    console.print(table)
