"""Parser for SPP (Souhrn Provedených Prací) Excel files."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from models import SPPItem, ColumnMapping


def col_letter_to_index(letter: str) -> int:
    """Convert column letter to 1-based index."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def to_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        cleaned = str(val).replace(",", ".").replace(" ", "").replace("\xa0", "").strip()
        if cleaned in ("", "-", "—"):
            return None
        # Handle percentage like "15%" or "0.15"
        if cleaned.endswith("%"):
            return float(cleaned[:-1])
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def parse_spp(
    filepath: str | Path,
    sheets: list[dict] | None = None,
    mapping: ColumnMapping | None = None,
    prefer_adaptive: bool = False,
) -> list[SPPItem]:
    """Parse an SPP Excel file into structured SPPItem objects.

    Args:
        filepath: Path to the Excel file (.xlsm or .xlsx).
        sheets: List of sheet configs [{"name": "...", "category_hint": "..."}].
                 If None, uses Chirana defaults.
        mapping: Column mapping. If None, uses Chirana defaults.

    Returns:
        List of SPPItem objects from all specified sheets.
    """
    if mapping is None:
        mapping = ColumnMapping(
            name="I",
            unit="K",
            quantity="L",
            price="M",       # price_per_unit
            total="N",
            percent_month="R",
            total_month="S",
            header_row=5,
            data_start_row=6,
        )

    auto_selected_sheets = sheets is None
    if sheets is None:
        sheets = _auto_select_spp_sheets(filepath)

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    all_items: list[SPPItem] = []
    month_hint = _extract_month_hint(str(filepath))

    for sheet_idx, sheet_cfg in enumerate(sheets):
        sheet_name = sheet_cfg["name"]
        if sheet_name not in wb.sheetnames:
            # Try partial match
            matched = [s for s in wb.sheetnames if sheet_name.lower() in s.lower()]
            if not matched:
                continue
            sheet_name = matched[0]

        ws = wb[sheet_name]
        adaptive = _detect_spp_mapping_from_headers(ws)
        if adaptive and (prefer_adaptive or not _mapping_has_enough_rows(ws, mapping)):
            active_mapping = adaptive
        else:
            active_mapping = mapping
        month_qty_col, month_total_col = _detect_month_columns(ws, active_mapping.header_row, month_hint)

        def get_cell(row_data: tuple, col_letter: str | None) -> Any:
            if not col_letter:
                return None
            idx = col_letter_to_index(col_letter) - 1
            if idx < len(row_data):
                cell = row_data[idx]
                try:
                    return cell.value if cell else None
                except AttributeError:
                    return None
            return None

        row_num = active_mapping.data_start_row - 1
        context_history: list[str] = []
        for row in ws.iter_rows(min_row=active_mapping.data_start_row):
            row_num += 1
            # Try to get actual row number from cell
            try:
                if row[0] and hasattr(row[0], 'row') and row[0].row:
                    row_num = row[0].row
            except (AttributeError, TypeError):
                pass

            name_val = get_cell(row, active_mapping.name)
            if not name_val or not str(name_val).strip():
                continue

            name_str = str(name_val).strip()
            quantity = to_float(get_cell(row, active_mapping.quantity))
            price_per = to_float(get_cell(row, active_mapping.price))
            total = to_float(get_cell(row, active_mapping.total))
            pct = to_float(get_cell(row, active_mapping.percent_month))
            total_m = to_float(get_cell(row, active_mapping.total_month))

            # Skip summary/total rows
            lower = name_str.lower()
            if any(kw in lower for kw in ["celkem", "součet", "total", "mezisoučet"]):
                continue

            # Many SPP formats carry current-month quantity/amount in separate month blocks.
            month_qty = to_float(get_cell(row, month_qty_col))
            month_total = to_float(get_cell(row, month_total_col))
            if month_qty is not None:
                quantity = month_qty
            if total_m is None and month_total is not None:
                total_m = month_total

            if _is_context_only_spp_row(name_str, quantity, total, total_m):
                if name_str not in context_history:
                    context_history.append(name_str)
                    context_history = context_history[-3:]
                continue

            effective_name = _compose_spp_item_name(name_str, context_history)

            item = SPPItem(
                row=(sheet_idx + 1) * 100000 + row_num,
                source_row=row_num,
                sheet=sheet_cfg.get("category_hint", sheet_name),
                name=effective_name,
                unit=str(get_cell(row, active_mapping.unit) or "").strip() or None,
                quantity=quantity,
                price_per_unit=price_per,
                total=total,
                percent_month=pct,
                total_month=total_m,
            )
            all_items.append(item)

    wb.close()
    # If explicit project-config sheets did not work, retry once with adaptive auto detection.
    if not all_items and not auto_selected_sheets:
        return parse_spp(filepath=filepath, sheets=None, mapping=mapping, prefer_adaptive=True)
    return all_items


def _is_spp_header_text(text: str) -> bool:
    normalized = text.lower()
    return (
        "název položky" in normalized
        or "nazev polozky" in normalized
        or "text položky" in normalized
        or "text polozky" in normalized
        or "popis položky" in normalized
        or "popis polozky" in normalized
        or normalized == "položka"
        or normalized == "polozka"
        or normalized == "popis"
    )


def _is_dimension_like_name(text: str) -> bool:
    normalized = str(text or "").strip()
    if not normalized:
        return False
    return bool(re.fullmatch(r"(?:[A-Za-z]{0,5}\s*)?\d{1,3}(?:[xX×]\d[\d,\.]*)+", normalized))


def _is_context_only_spp_row(
    name: str,
    quantity: float | None,
    total: float | None,
    total_month: float | None,
) -> bool:
    lower = str(name or "").strip().lower()
    if not lower:
        return True
    if _is_dimension_like_name(lower):
        return False
    if quantity is not None or total not in (None, 0) or total_month not in (None, 0):
        return False
    return True


def _compose_spp_item_name(name: str, context_history: list[str]) -> str:
    if not context_history:
        return name
    if _is_dimension_like_name(name) or len(name.strip()) <= 24:
        merged = " | ".join(context_history[-2:] + [name])
        return merged[:300]
    return name


def _looks_like_month_label(text: str) -> bool:
    normalized = str(text or "").strip().lower()
    if not normalized:
        return False
    month_tokens = (
        "leden", "únor", "unor", "březen", "brezen", "duben", "květen", "kveten",
        "červen", "cerven", "červenec", "cervenec", "srpen", "září", "zari",
        "říjen", "rijen", "listopad", "prosinec", "january", "february", "march",
        "april", "may", "june", "july", "august", "september", "october",
        "november", "december",
    )
    return any(token in normalized for token in month_tokens) or bool(re.search(r"20\d{2}", normalized))


def filter_spp_by_month(items: list[SPPItem]) -> list[SPPItem]:
    """Keep only SPP items that have actual work for the period.

    Business rule: active month rows are only those with percent_month > 0.
    """
    # Preserve legacy Chirana behavior when percent_month is available:
    # active row <=> percent_month > 0.
    has_percent_data = any(item.percent_month is not None for item in items)
    result: list[SPPItem] = []
    for item in items:
        pct = item.percent_month or 0
        total_month = item.total_month or 0
        if has_percent_data:
            if pct > 0:
                result.append(item)
        else:
            if total_month > 0:
                result.append(item)
    return result


def get_spp_preview(
    filepath: str | Path,
    sheet_name: str | None = None,
    num_rows: int = 10,
) -> list[list[Any]]:
    """Get first N rows from an SPP sheet for preview."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(max_row=num_rows)):
        rows.append([cell.value for cell in row])
    wb.close()
    return rows


def _extract_month_hint(path_value: str) -> str:
    low = path_value.lower()
    month_aliases = {
        "leden": "leden",
        "led": "leden",
        "january": "leden",
        "unor": "unor",
        "únor": "unor",
        "feb": "unor",
        "february": "unor",
        "brezen": "brezen",
        "březen": "brezen",
        "mar": "brezen",
        "march": "brezen",
        "duben": "duben",
        "apr": "duben",
        "kveten": "kveten",
        "květen": "kveten",
        "may": "kveten",
        "cerven": "cerven",
        "červen": "cerven",
        "jun": "cerven",
        "cervenec": "cervenec",
        "červenec": "cervenec",
        "jul": "cervenec",
        "srpen": "srpen",
        "aug": "srpen",
        "zari": "zari",
        "září": "zari",
        "sep": "zari",
        "rijen": "rijen",
        "říjen": "rijen",
        "oct": "rijen",
        "listopad": "listopad",
        "nov": "listopad",
        "prosinec": "prosinec",
        "dec": "prosinec",
    }
    for alias, normalized in month_aliases.items():
        if alias in low:
            return normalized
    return ""


def _detect_month_columns(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    month_hint: str,
) -> tuple[str | None, str | None]:
    scan_from = max(1, header_row - 3)
    scan_to = min(ws.max_row, header_row + 8)
    max_scan_col = min(260, ws.max_column)
    candidates: list[tuple[int, int, int, int | None, int | None]] = []

    for month_row in range(scan_from, scan_to + 1):
        descriptor_rows = range(
            month_row + 1,
            min(ws.max_row, max(header_row + 2, month_row + 1), month_row + 3) + 1,
        )
        for col in range(1, max_scan_col + 1):
            month_text = str(ws.cell(month_row, col).value or "").strip().lower()
            if not _looks_like_month_label(month_text):
                continue

            score = 25
            if month_hint and month_hint in month_text:
                score += 100

            qty_col = None
            total_col = None
            descriptor_row_for_hits = month_row + 1
            for desc_row in descriptor_rows:
                for probe in range(col, min(col + 3, max_scan_col) + 1):
                    sub_text = str(ws.cell(desc_row, probe).value or "").strip().lower()
                    if any(k in sub_text for k in ["množ", "mnoz", "qty", "jm"]):
                        qty_col = probe
                        descriptor_row_for_hits = desc_row
                    if any(k in sub_text for k in ["celkem", "cena", "částka", "castka"]):
                        total_col = probe
                        descriptor_row_for_hits = desc_row

            if qty_col:
                score += 20
            if total_col:
                score += 35
            if qty_col and total_col and total_col >= qty_col:
                score += 10
            qty_hits = _count_nonzero_numeric_cells(ws, qty_col, descriptor_row_for_hits + 1) if qty_col else 0
            total_hits = _count_nonzero_numeric_cells(ws, total_col, descriptor_row_for_hits + 1) if total_col else 0
            score += min(qty_hits, 10) * 3
            score += min(total_hits, 10) * 4
            if score > 25:
                candidates.append((score, total_hits, qty_hits, qty_col, total_col))

    if not candidates:
        return None, None

    candidates.sort(key=lambda x: (x[0], x[1], x[2], x[4] or 0, x[3] or 0), reverse=True)
    _, _, _, qty_col, total_col = candidates[0]
    qty_letter = openpyxl.utils.get_column_letter(qty_col) if qty_col else None
    total_letter = openpyxl.utils.get_column_letter(total_col) if total_col else None
    return qty_letter, total_letter


def _auto_select_spp_sheets(filepath: str | Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    result: list[dict[str, str]] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            hit = _detect_spp_mapping_from_headers(ws) is not None
            if not hit:
                for row_idx in range(1, 16):
                    row_values = [ws.cell(row_idx, c).value for c in range(1, min(40, ws.max_column) + 1)]
                    row_text = " | ".join(str(v) for v in row_values if v not in (None, "")).lower()
                    if "výkaz výměr" in row_text or "vykaz vymer" in row_text:
                        hit = True
                    if "množství" in row_text or "mnozstvi" in row_text:
                        if any(key in row_text for key in ["cena celkem", "celková cena", "celkova cena", "jednotka", "mj"]):
                            hit = True
                    if hit:
                        break
            if hit:
                normalized = _normalize_sheet_hint(name)
                result.append({"name": name, "category_hint": normalized})
    finally:
        wb.close()
    if result:
        return result
    # Safe fallback to historical defaults if no headers were recognized.
    return [
        {"name": "Fakturace SoD - ZTI", "category_hint": "ZTI"},
        {"name": "Fakturace SoD - ÚT", "category_hint": "UT"},
    ]


def _normalize_sheet_hint(sheet_name: str) -> str:
    low = sheet_name.lower()
    if "zti" in low:
        return "ZTI"
    if "ut" in low or "út" in low or "vytáp" in low:
        return "UT"
    if "kanal" in low:
        return "KAN"
    if "vod" in low:
        return "VOD"
    if "vzt" in low:
        return "VZT"
    m = re.sub(r"[^A-Za-z0-9]+", " ", sheet_name).strip()
    return m[:20] if m else sheet_name[:20]


def _detect_spp_mapping_from_headers(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> ColumnMapping | None:
    max_col = min(ws.max_column, 120)
    header_row = None
    name_col = None
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        normalized = []
        for c in range(1, max_col + 1):
            val = ws.cell(row_idx, c).value
            normalized.append((c, str(val).strip().lower() if val not in (None, "") else ""))
        best_name_score = 0
        for c, text in normalized:
            score = 0
            if "popis položky" in text or "popis polozky" in text:
                score = 5
            elif "název položky" in text or "nazev polozky" in text:
                score = 5
            elif text == "popis":
                score = 4
            elif text == "položka" or text == "polozka":
                score = 3
            elif _is_spp_header_text(text):
                score = 1
            if score > best_name_score:
                best_name_score = score
                header_row = row_idx
                name_col = c
        if header_row:
            break
    if not header_row or not name_col:
        return None

    def _find_col(keywords: tuple[str, ...], default: int | None = None) -> int | None:
        for c in range(1, max_col + 1):
            text = str(ws.cell(header_row, c).value or "").strip().lower()
            if any(k in text for k in keywords):
                return c
        return default

    unit_col = _find_col(("mj", "jednotk", "unit"), default=name_col + 1)
    qty_col = _find_col(("množství", "mnozstvi", "qty"), default=name_col + 2)
    price_col = _find_col(("j.cena", "jedn.cena", "dodávka", "dodavka", "price"), default=name_col + 3)
    total_col = _find_col(("cena s dph", "cena celkem", "celk.", "celková cena", "celkova cena", "cena", "total"), default=name_col + 4)
    row_number_col = _find_col(("číslo položky", "cislo polozky", "č.p.", "č.p", "poř.", "por.", "kód", "kod"), default=max(1, name_col - 1))
    data_start_row = _detect_spp_data_start_row(ws, header_row, name_col, qty_col, total_col, row_number_col)
    return ColumnMapping(
        row_number=openpyxl.utils.get_column_letter(row_number_col) if row_number_col else None,
        name=openpyxl.utils.get_column_letter(name_col),
        unit=openpyxl.utils.get_column_letter(unit_col) if unit_col else None,
        quantity=openpyxl.utils.get_column_letter(qty_col) if qty_col else None,
        price=openpyxl.utils.get_column_letter(price_col) if price_col else None,
        total=openpyxl.utils.get_column_letter(total_col) if total_col else None,
        percent_month=None,
        total_month=None,
        header_row=header_row,
        data_start_row=data_start_row,
    )


def _detect_spp_data_start_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    name_col: int,
    qty_col: int | None,
    total_col: int | None,
    row_number_col: int | None,
) -> int:
    scan_to = min(ws.max_row, header_row + 40)
    for row_idx in range(header_row + 1, scan_to + 1):
        name_val = ws.cell(row_idx, name_col).value
        if name_val in (None, "") or not str(name_val).strip():
            continue
        name_text = str(name_val).strip().lower()
        if any(skip in name_text for skip in ["množství", "mnozstvi", "cena celkem", "celková cena", "poznámka", "poznamka"]):
            continue

        row_code = str(ws.cell(row_idx, row_number_col).value or "").strip() if row_number_col else ""
        qty_val = ws.cell(row_idx, qty_col).value if qty_col else None
        total_val = ws.cell(row_idx, total_col).value if total_col else None
        if row_code or to_float(qty_val) is not None or to_float(total_val) is not None:
            return row_idx
    return header_row + 1


def _count_nonzero_numeric_cells(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    col_idx: int | None,
    start_row: int,
    window: int = 60,
) -> int:
    if not col_idx:
        return 0
    hits = 0
    end_row = min(ws.max_row, start_row + window)
    for row_idx in range(start_row, end_row + 1):
        value = to_float(ws.cell(row_idx, col_idx).value)
        if value and abs(value) > 0:
            hits += 1
    return hits


def _mapping_has_enough_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    mapping: ColumnMapping,
    min_rows: int = 5,
) -> bool:
    if not mapping.name:
        return False
    idx = col_letter_to_index(mapping.name) - 1
    non_empty = 0
    for row in ws.iter_rows(min_row=max(1, mapping.data_start_row), max_row=min(ws.max_row, mapping.data_start_row + 60)):
        if idx < len(row):
            val = row[idx].value
            if val not in (None, "") and str(val).strip():
                non_empty += 1
                if non_empty >= min_rows:
                    return True
    return False
