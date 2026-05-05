"""Localhost web app for Domostav AI demo."""

from __future__ import annotations

import importlib
import hashlib
import json
import os
import re
import time
from datetime import datetime
from pathlib import Path
import uuid

import openpyxl
from flask import Flask, g, jsonify, redirect, render_template, request, send_from_directory, session, url_for

from auth import (
    ROLE_DEFINITIONS,
    authenticate,
    can_access_module,
    can_manage_users,
    current_user,
    forbidden_response,
    login_required,
    login_user,
    logout_user,
    require_roles,
    role_exists,
    role_label,
    store as auth_store,
)
from config import settings
from config.settings import DATA_DIR, OUTPUT_DIR, PROJECT_ROOT
from db import apply_schemas
from llm.client import ClaudeClient
from models import ColumnMapping, WriteoffRecommendation
from parsers.mapping_engine import auto_detect_mapping
from services.pipeline_service import mapping_to_dict
from storage.session_store import SessionStore

# Create tables for both schemas before any route or blueprint touches the DB.
apply_schemas()

app = Flask(__name__, static_folder="design", static_url_path="")
app.config["JSON_AS_ASCII"] = False
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024  # 64 MB per upload batch
app.jinja_env.auto_reload = True
app.secret_key = (
    getattr(settings, "APP_SECRET_KEY", "").strip()
    or hashlib.sha256((os.environ.get("DATABASE_URL") or str(PROJECT_ROOT)).encode("utf-8")).hexdigest()
)

from mzdovy import blueprint as mzdovy_blueprint

app.register_blueprint(mzdovy_blueprint)

UPLOAD_DIR = DATA_DIR / "uploads"
store = SessionStore()
_DEBUG_LOG_PATH = "/Users/dmytriivezerian/Desktop/Domostav x Fajnwork/.cursor/debug-f07731.log"

DEFAULT_SPP_PROJECTS = [
    ("bd-makovska", "BD Makovska"),
    ("bd-ohrada", "BD Ohrada"),
    ("chirana", "CHIRANA (NOVECON)"),
    ("odkolek", "Odkolek"),
    ("rezidence-nad-vltavou", "Rezidence nad Vltavou"),
]


def _ensure_default_spp_projects() -> None:
    for code, name in DEFAULT_SPP_PROJECTS:
        store.ensure_project(code, name)


_ensure_default_spp_projects()


def _project_code_from_name(value: str) -> str:
    text = value.strip().lower()
    text = (
        text.replace("á", "a").replace("č", "c").replace("ď", "d")
        .replace("é", "e").replace("ě", "e").replace("í", "i")
        .replace("ň", "n").replace("ó", "o").replace("ř", "r")
        .replace("š", "s").replace("ť", "t").replace("ú", "u")
        .replace("ů", "u").replace("ý", "y").replace("ž", "z")
    )
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or f"projekt-{int(time.time())}"


@app.context_processor
def inject_auth_context():
    user = current_user()
    return {
        "current_user": user,
        "role_label": role_label,
        "can_access_spp": can_access_module(user, "spp"),
        "can_access_mzdovy": can_access_module(user, "mzdovy"),
        "can_manage_users": can_manage_users(user),
    }


@app.before_request
def enforce_auth():
    user = current_user()
    public_paths = {
        "/health",
        "/api/version",
        "/login",
        "/logout",
        "/setup",
    }
    if request.path in public_paths or request.path.startswith("/mzdovy/static/"):
        return None
    if not auth_store.has_users() and request.path != "/setup":
        return redirect(url_for("setup_page"))
    if request.path.startswith("/mzdovy"):
        return None
    if request.path == "/users" or request.path.startswith("/users/"):
        if not user:
            return redirect(url_for("login_page", next=request.path))
        if not can_manage_users(user):
            return render_template("access_denied.html", title="Přístup odepřen", message="Na správu uživatelů nemáte oprávnění."), 403
        return None
    if request.path == "/login":
        return None
    if request.path.startswith("/api/me"):
        if not user:
            return jsonify({"error": "Přihlaste se prosím."}), 401
        return None
    if request.path == "/" or request.path.startswith("/api/") or request.path.startswith("/output/"):
        if not user:
            return redirect(url_for("login_page", next=request.path)) if not request.path.startswith("/api/") else (jsonify({"error": "Přihlaste se prosím."}), 401)
        if not can_access_module(user, "spp"):
            if request.path == "/" and can_access_module(user, "mzdovy"):
                return redirect(url_for("mzdovy.wizard_new"))
            return (
                (jsonify({"error": "Na Modul 1 nemáte přístup."}), 403)
                if request.path.startswith("/api/")
                else (render_template("access_denied.html", title="Přístup odepřen", message="Na Modul 1 nemáte přístup."), 403)
            )
    return None


def _debug_log(hypothesis_id: str, location: str, message: str, data: dict) -> None:
    payload = {
        "sessionId": "f07731",
        "runId": "initial",
        "hypothesisId": hypothesis_id,
        "location": location,
        "message": message,
        "data": data,
        "timestamp": int(time.time() * 1000),
    }
    try:
        with open(_DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        pass


def _ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _bool(value: str | None) -> bool:
    if value is None:
        return False
    return value.lower() in ("1", "true", "yes", "y", "on")


def _save_uploaded_file(file_obj, session_dir: Path) -> str:
    safe_name = file_obj.filename or f"upload_{uuid.uuid4().hex}"
    path = session_dir / safe_name
    file_obj.save(path)
    return str(path)


def _override_item_key(article: str | None, inventory_row: int | None) -> str | None:
    normalized = (article or "").strip().upper()
    if normalized:
        return f"ARTICLE:{normalized}"
    if inventory_row is not None:
        return f"ROW:{int(inventory_row)}"
    return None


def _load_pipeline_module():
    import output.excel_generator as excel_gen
    import services.pipeline_service as pipeline_mod

    importlib.reload(excel_gen)
    importlib.reload(pipeline_mod)
    return pipeline_mod


def _run_pipeline(
    *,
    project_code: str,
    spp_path: str,
    inventory_path: str,
    period_month: str,
    no_ai: bool,
    auto_map: bool,
    spp_mapping_override: dict | None,
    inv_mapping_override: dict | None,
    project_prompt: str,
    rules_path: str | None,
    nomenclature_path: str | None,
    nf45_path: str | None,
    generate_excel: bool,
):
    effective_overrides = store.get_effective_overrides(project_code)
    alias_rules = store.get_effective_rules(project_code, "alias")
    category_rules = store.get_effective_rules(project_code, "category")
    alias_map: dict[str, str] = {}
    for rule in alias_rules:
        val = rule.get("rule_value") or {}
        alias = (val.get("alias") or rule.get("rule_key") or "").strip()
        canonical = (val.get("canonical") or "").strip()
        if alias and canonical:
            alias_map[alias] = canonical

    # region agent log
    _debug_log(
        "H3",
        "webapp.py:_run_pipeline",
        "Pipeline input context",
        {
            "project_code": project_code,
            "effective_override_count": len(effective_overrides),
            "effective_override_sample": list(effective_overrides.keys())[:8],
            "rules_path_present": bool(rules_path),
            "nomenclature_present": bool(nomenclature_path),
        },
    )
    # endregion

    pipeline_mod = _load_pipeline_module()
    return pipeline_mod.run_analysis_pipeline(
        project=project_code,
        spp_path=spp_path,
        inventory_path=inventory_path,
        period_month=period_month,
        api_key=settings.OPENAI_API_KEY or settings.ANTHROPIC_API_KEY,
        no_ai=no_ai,
        auto_map=auto_map,
        spp_mapping_override=spp_mapping_override,
        inv_mapping_override=inv_mapping_override,
        project_prompt_override=project_prompt,
        rules_xlsm_path=rules_path,
        nomenclature_path=nomenclature_path,
        nf45_path=nf45_path,
        overrides=effective_overrides,
        alias_map=alias_map,
        category_rules=category_rules,
        generate_excel=generate_excel,
        include_export_artifacts=not generate_excel,
    )


def _draft_artifacts_path(spp_path: str) -> Path:
    return Path(spp_path).parent / "analysis_artifacts.json"


def _write_draft_artifacts(spp_path: str, result: dict) -> None:
    payload = {
        "result_snapshot": {k: v for k, v in result.items() if k != "export_artifacts"},
        "export_artifacts": result.get("export_artifacts") or {},
    }
    _draft_artifacts_path(spp_path).write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _read_draft_artifacts(spp_path: str) -> dict | None:
    path = _draft_artifacts_path(spp_path)
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def _strip_export_artifacts(result: dict) -> dict:
    return {k: v for k, v in result.items() if k != "export_artifacts"}


def _finalize_from_cached_artifacts(draft: dict) -> dict[str, object]:
    cached = _read_draft_artifacts(draft["spp_path"])
    if not cached:
        raise FileNotFoundError("Cached analysis artifacts not found")

    export_artifacts = cached.get("export_artifacts") or {}
    recommendations = [
        WriteoffRecommendation(**item)
        for item in (export_artifacts.get("recommendations") or [])
    ]
    output_path = str(OUTPUT_DIR / f"analysis_{draft['project_code']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    import output.excel_generator as excel_gen

    excel_gen.generate_output(
        source_path=draft["inventory_path"],
        output_path=output_path,
        recommendations=recommendations,
        data_start_row=int(export_artifacts.get("data_start_row") or 12),
        sheet_name=export_artifacts.get("sheet_name"),
        summary=export_artifacts.get("summary") or {},
        spp_coverage=export_artifacts.get("spp_coverage") or [],
        header_row=export_artifacts.get("header_row"),
    )
    result = dict(cached.get("result_snapshot") or {})
    result["output_path"] = output_path
    return result


def _validate_upload_slots(spp, inventory, nf45, rules_file) -> str | None:
    """Filename checks are intentionally disabled to avoid false rejects."""
    _ = (spp, inventory, nf45, rules_file)
    return None


def _default_mapping(file_type: str) -> ColumnMapping:
    """Fallback mapping for known Chirana templates when AI is unavailable."""
    if file_type == "spp":
        return ColumnMapping(
            name="I",
            unit="K",
            quantity="L",
            price="M",
            total="N",
            percent_month="R",
            total_month="S",
            header_row=5,
            data_start_row=6,
        )
    return ColumnMapping(
        number="B",
        article="D",
        name="F",
        deviation="K",
        quantity="N",
        quantity_accounting="Q",
        unit="T",
        price="V",
        header_row=11,
        data_start_row=12,
    )


def _find_default_rules_file() -> str | None:
    workspace = PROJECT_ROOT.parent.parent
    candidates = list(workspace.glob("SPP_Chirana_02_26_*Topeni+Kanalizace.xlsm"))
    return str(candidates[0]) if candidates else None


def _find_default_nomenclature_file() -> str | None:
    workspace = PROJECT_ROOT.parent.parent
    candidates = list(workspace.glob("*номенклатури*.xlsx"))
    return str(candidates[0]) if candidates else None


def _detect_training_file_kind(path: Path, filename: str) -> str:
    low = filename.lower()
    if any(k in low for k in ["spp", "rozpo", "fakturace sod", "soupis prac", "výkaz výměr", "vykaz vymer"]):
        return "spp"
    if any(k in low for k in ["invent", "nf-30", "sklad", "zásob", "zasob", "fakturace_"]):
        return "inventory"
    if any(k in low for k in ["nf-45", "spisani", "списан"]):
        return "writeoff"
    if any(k in low for k in ["pravid", "rules", "topeni+kanalizace"]):
        return "rules"
    if any(k in low for k in ["nomen", "номенклат"]):
        return "nomenclature"

    # Lightweight header-based fallback for xlsx/xlsm.
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
        return "unknown"
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        ws = wb.active
        text_parts: list[str] = []
        max_col = min(ws.max_column, 40)
        max_row = min(ws.max_row, 40)
        for r in range(1, max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
            text_parts.extend(str(v) for v in vals if v not in (None, ""))
        wb.close()
        header_blob = " | ".join(text_parts).lower()
        if "název položky" in header_blob or "nazev polozky" in header_blob:
            return "spp"
        if "invent" in header_blob or "odchyl" in header_blob or "deviation" in header_blob:
            return "inventory"
        if "soupis prac" in header_blob and "pč" in header_blob and "kód" in header_blob:
            return "inventory"
        if "nomenkl" in header_blob:
            return "nomenclature"
    except Exception:
        return "unknown"
    return "unknown"


def _role_options() -> list[dict[str, str]]:
    return [{"value": key, "label": value["label"]} for key, value in ROLE_DEFINITIONS.items()]


def _safe_next_url_for_user(user: dict[str, object], next_url: str) -> str | None:
    next_url = (next_url or "").strip()
    if not next_url.startswith("/") or next_url.startswith("//"):
        return None
    if next_url == "/" or next_url.startswith("/api/") or next_url.startswith("/output/"):
        return next_url if can_access_module(user, "spp") else None
    if next_url.startswith("/mzdovy"):
        return next_url if can_access_module(user, "mzdovy") else None
    if next_url == "/users" or next_url.startswith("/users/"):
        return next_url if can_manage_users(user) else None
    return None


@app.route("/login", methods=["GET", "POST"])
def login_page():
    if not auth_store.has_users():
        return redirect(url_for("setup_page"))
    existing_user = current_user()
    if existing_user:
        if can_access_module(existing_user, "spp"):
            return redirect(url_for("root"))
        if can_access_module(existing_user, "mzdovy"):
            return redirect(url_for("mzdovy.wizard_new"))
        return redirect(url_for("logout_page"))
    error = ""
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        user = authenticate(username, password)
        if user:
            login_user(user)
            next_url = _safe_next_url_for_user(user, request.args.get("next", ""))
            if next_url:
                return redirect(next_url)
            if can_access_module(user, "spp"):
                return redirect(url_for("root"))
            if can_access_module(user, "mzdovy"):
                return redirect(url_for("mzdovy.wizard_new"))
            error = "Účet nemá přiřazený žádný dostupný modul."
        else:
            error = "Neplatné přihlašovací údaje."
    return render_template("login.html", title="Přihlášení", error=error)


@app.route("/setup", methods=["GET", "POST"])
def setup_page():
    if auth_store.has_users():
        return redirect(url_for("login_page"))
    error = ""
    if request.method == "POST":
        full_name = (request.form.get("full_name") or "").strip()
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        password_confirm = request.form.get("password_confirm") or ""
        if not full_name or not username or not password:
            error = "Vyplňte prosím všechna pole."
        elif password != password_confirm:
            error = "Hesla se neshodují."
        elif len(password) < 8:
            error = "Heslo musí mít alespoň 8 znaků."
        else:
            try:
                auth_store.create_user(
                    username=username,
                    full_name=full_name,
                    password=password,
                    role="super_admin",
                )
                user = authenticate(username, password)
                if user:
                    login_user(user)
                return redirect(url_for("users_page"))
            except Exception:
                error = "Tento uživatel už pravděpodobně existuje."
    return render_template("setup.html", title="První nastavení", error=error)


@app.get("/logout")
def logout_page():
    logout_user()
    return redirect(url_for("login_page"))


@app.get("/api/me")
def api_me():
    user = current_user()
    if not user:
        return jsonify({"error": "Přihlaste se prosím."}), 401
    return jsonify(
        {
            "user": {
                "id": user["id"],
                "username": user["username"],
                "full_name": user["full_name"],
                "role": user["role"],
                "role_label": role_label(user["role"]),
                "modules": sorted(ROLE_DEFINITIONS[user["role"]]["modules"]),
                "can_manage_users": can_manage_users(user),
            }
        }
    )


@app.get("/users")
@require_roles("super_admin", "admin")
def users_page():
    users = auth_store.list_users()
    for user in users:
        user["role_label"] = role_label(user.get("role"))
    return render_template(
        "users.html",
        title="Uživatelé",
        eyebrow="Administrace",
        page_title="Uživatelé platformy",
        page_description="Správa účtů, rolí a přístupů ve stejném prostředí jako oba moduly.",
        active_page="users",
        users=users,
        roles=_role_options(),
        error="",
    )


@app.post("/users/create")
@require_roles("super_admin", "admin")
def create_user_page():
    full_name = (request.form.get("full_name") or "").strip()
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    role = (request.form.get("role") or "").strip()
    error = ""
    if not full_name or not username or not password:
        error = "Vyplňte prosím všechna pole pro nového uživatele."
    elif len(password) < 8:
        error = "Dočasné heslo musí mít alespoň 8 znaků."
    elif not role_exists(role):
        error = "Vyberte prosím platnou roli."
    else:
        try:
            auth_store.create_user(username=username, full_name=full_name, password=password, role=role)
            return redirect(url_for("users_page"))
        except Exception:
            error = "Tento uživatel už pravděpodobně existuje."
    users = auth_store.list_users()
    for user in users:
        user["role_label"] = role_label(user.get("role"))
    return render_template(
        "users.html",
        title="Uživatelé",
        eyebrow="Administrace",
        page_title="Uživatelé platformy",
        page_description="Správa účtů, rolí a přístupů ve stejném prostředí jako oba moduly.",
        active_page="users",
        users=users,
        roles=_role_options(),
        error=error,
    ), 400


@app.post("/users/<int:user_id>/update")
@require_roles("super_admin", "admin")
def update_user_page(user_id: int):
    role = (request.form.get("role") or "").strip()
    is_active = (request.form.get("is_active") or "").strip().lower() == "true"
    password = (request.form.get("password") or "").strip()
    if not role_exists(role):
        return render_template(
            "access_denied.html",
            title="Neplatná role",
            message="Vybraná role není platná.",
        ), 400
    auth_store.update_user(user_id, role=role, is_active=is_active, password=password or None)
    if session.get("auth_user_id") == user_id and not is_active:
        logout_user()
        return redirect(url_for("login_page"))
    return redirect(url_for("users_page"))


@app.post("/users/<int:user_id>/delete")
@require_roles("super_admin", "admin")
def delete_user_page(user_id: int):
    if session.get("auth_user_id") == user_id:
        users = auth_store.list_users()
        for user in users:
            user["role_label"] = role_label(user.get("role"))
        return render_template(
            "users.html",
            title="Uživatelé",
            eyebrow="Administrace",
            page_title="Uživatelé platformy",
            page_description="Správa účtů, rolí a přístupů ve stejném prostředí jako oba moduly.",
            active_page="users",
            users=users,
            roles=_role_options(),
            error="Vlastní účet nelze smazat během aktivní relace.",
        ), 400
    auth_store.delete_user(user_id)
    return redirect(url_for("users_page"))


@app.get("/projects/new")
@login_required
def new_project_page():
    if not can_access_module(current_user(), "spp"):
        return forbidden_response("Na zakládání SPP projektů nemáte přístup.")
    projects = store.list_projects(include_archived=True)
    return render_template(
        "project_new.html",
        title="Nový projekt",
        eyebrow="Projekty",
        page_title="Nový projekt",
        page_description="Samostatný objekt pro Module 1, jeho pravidla, analýzy a historii.",
        active_page="project_new",
        projects=projects,
        error="",
    )


@app.post("/projects/create")
@login_required
def create_project_page():
    if not can_access_module(current_user(), "spp"):
        return forbidden_response("Na zakládání SPP projektů nemáte přístup.")
    name = (request.form.get("name") or "").strip()
    code = (request.form.get("code") or "").strip().lower()
    prompt = request.form.get("prompt") or ""
    if not name:
        return render_template(
            "project_new.html",
            title="Nový projekt",
            eyebrow="Projekty",
            page_title="Nový projekt",
            page_description="Samostatný objekt pro Module 1, jeho pravidla, analýzy a historii.",
            active_page="project_new",
            projects=store.list_projects(include_archived=True),
            error="Zadejte prosím název projektu.",
        ), 400
    store.ensure_project(code or _project_code_from_name(name), name, prompt)
    return redirect(url_for("root"))


@app.post("/projects/<project_code>/archive")
@login_required
def archive_project_page(project_code: str):
    if not can_access_module(current_user(), "spp"):
        return forbidden_response("Na správu SPP projektů nemáte přístup.")
    store.archive_project(project_code)
    return redirect(url_for("new_project_page"))


@app.post("/projects/<project_code>/restore")
@login_required
def restore_project_page(project_code: str):
    if not can_access_module(current_user(), "spp"):
        return forbidden_response("Na správu SPP projektů nemáte přístup.")
    store.restore_project(project_code)
    return redirect(url_for("new_project_page"))


@app.post("/projects/<project_code>/delete")
@login_required
def delete_project_page(project_code: str):
    if not can_access_module(current_user(), "spp"):
        return forbidden_response("Na správu SPP projektů nemáte přístup.")
    deleted = store.delete_project_if_empty(project_code)
    if deleted:
        return redirect(url_for("new_project_page"))
    projects = store.list_projects(include_archived=True)
    return render_template(
        "project_new.html",
        title="Nový projekt",
        eyebrow="Projekty",
        page_title="Nový projekt",
        page_description="Samostatný objekt pro Module 1, jeho pravidla, analýzy a historii.",
        active_page="project_new",
        projects=projects,
        error="Projekt už má analýzy nebo pravidla. Kvůli historii jej archivujte místo mazání.",
    ), 400


@app.get("/")
def root():
    return send_from_directory("design", "dashboard.html")


@app.get("/health")
def health():
    return jsonify({"ok": True, "service": "domostav-ai-web"})


@app.get("/api/version")
def api_version():
    """Expose active runtime components for quick verification."""
    import output.excel_generator as excel_gen
    return jsonify(
        {
            "ok": True,
            "excel_generator_file": str(Path(excel_gen.__file__).resolve()),
            "ai_columns": [name for name, _ in excel_gen.AI_COLUMNS],
            "supports_spp_coverage": hasattr(excel_gen, "_add_spp_coverage_sheet"),
        }
    )


@app.get("/api/projects")
def list_projects():
    return jsonify({"projects": store.list_projects()})


@app.post("/api/projects")
def create_project():
    payload = request.get_json(silent=True) or {}
    code = (payload.get("code") or "").strip().lower()
    name = (payload.get("name") or "").strip()
    prompt = payload.get("prompt") or ""
    if not code:
        return jsonify({"error": "code is required"}), 400
    store.ensure_project(code, name or code.title(), prompt)
    return jsonify({"ok": True})


@app.get("/api/sessions")
def list_sessions():
    project = request.args.get("project")
    month = request.args.get("month")
    sessions = store.list_sessions(project_code=project, period_month=month)
    for s in sessions:
        output_path = s.get("output_path") or ""
        p = Path(output_path)
        s["output_exists"] = p.exists()
        s["output_filename"] = p.name if p.name else None
    return jsonify({"sessions": sessions})


@app.get("/api/dashboard")
def dashboard_stats():
    project = request.args.get("project", "chirana")
    month = request.args.get("month")
    sessions = store.list_sessions(project_code=project, period_month=month)
    latest = sessions[0] if sessions else None
    monthly = {}
    for s in sessions:
        key = s["period_month"]
        monthly.setdefault(
            key,
            {
                "count": 0,
                "ok": 0,
                "warning": 0,
                "red_flag": 0,
            },
        )
        monthly[key]["count"] += 1
        kpi = s["stats"].get("kpis", {})
        monthly[key]["ok"] += int(kpi.get("ok", 0))
        monthly[key]["warning"] += int(kpi.get("warning", 0))
        monthly[key]["red_flag"] += int(kpi.get("red_flag", 0))

    return jsonify(
        {
            "project": project,
            "latest": latest,
            "monthly": monthly,
            "sessions_count": len(sessions),
        }
    )


@app.post("/api/overrides")
def save_override():
    payload = request.get_json(silent=True) or {}
    project = payload.get("project_code", "chirana")
    scope = (payload.get("scope") or "project").strip().lower()
    action = (payload.get("action") or "save").strip().lower()
    article = (payload.get("article") or "").strip().upper()
    inventory_row = payload.get("inventory_row")
    item_key = (payload.get("item_key") or "").strip()
    rows = payload.get("spp_rows") or []
    reason = payload.get("reason") or ""
    if not item_key:
        item_key = _override_item_key(article, inventory_row)
    if action == "delete":
        removed: list[str] = []
        if scope in {"system", "both"}:
            if store.delete_scoped_override("system", "global", item_key):
                removed.append("system")
        if scope in {"project", "both"}:
            if store.delete_scoped_override("project", project, item_key):
                removed.append("project")
        status = store.get_override_status(project, item_key)
        # region agent log
        _debug_log(
            "H4",
            "webapp.py:/api/overrides",
            "Override delete requested",
            {
                "project": project,
                "scope": scope,
                "item_key": item_key,
                "removed_scopes": removed,
                "status_after_delete": status,
            },
        )
        # endregion
        return jsonify({"ok": True, "action": "delete", "removed_scopes": removed, "status": status})

    if not item_key or not isinstance(rows, list):
        return jsonify({"error": "item_key/article and spp_rows are required"}), 400

    normalized_rows = [int(x) for x in rows]
    saved: list[str] = []
    if scope in {"system", "both"}:
        store.save_scoped_override("system", "global", item_key, normalized_rows, reason)
        saved.append("system")
    if scope in {"project", "both"}:
        store.save_scoped_override("project", project, item_key, normalized_rows, reason)
        saved.append("project")
    status = store.get_override_status(project, item_key)
    # region agent log
    _debug_log(
        "H4",
        "webapp.py:/api/overrides",
        "Override save requested",
        {
            "project": project,
            "scope": scope,
            "item_key": item_key,
            "saved_scopes": saved,
            "status_after_save": status,
        },
    )
    # endregion
    return jsonify({"ok": True, "action": "save", "saved_scopes": saved, "status": status})


@app.post("/api/overrides/bulk")
def save_overrides_bulk():
    payload = request.get_json(silent=True) or {}
    project = (payload.get("project_code") or "chirana").strip().lower()
    scope = (payload.get("scope") or "project").strip().lower()
    rows = payload.get("rows") or []
    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "rows array is required"}), 400

    saved_count = 0
    failed: list[dict] = []
    for idx, row in enumerate(rows):
        try:
            article = (row.get("article") or "").strip().upper()
            inventory_row = row.get("inventory_row")
            item_key = (row.get("item_key") or "").strip() or _override_item_key(article, inventory_row)
            raw_spp_rows = row.get("spp_rows")
            if raw_spp_rows is None:
                raw_single = row.get("spp_row")
                raw_spp_rows = [raw_single] if raw_single not in (None, "", 0) else []
            if not isinstance(raw_spp_rows, list):
                raise ValueError("spp_rows must be an array")
            normalized_rows: list[int] = []
            for x in raw_spp_rows:
                try:
                    value = int(x)
                except (TypeError, ValueError):
                    continue
                if value > 0:
                    normalized_rows.append(value)
            reason = row.get("reason") or ""
            if not item_key or not normalized_rows:
                raise ValueError("item_key/article and spp_rows are required")
            if scope in {"system", "both"}:
                store.save_scoped_override("system", "global", item_key, normalized_rows, reason)
            if scope in {"project", "both"}:
                store.save_scoped_override("project", project, item_key, normalized_rows, reason)
            saved_count += 1
        except Exception as exc:
            failed.append({"index": idx, "error": str(exc)})
    # region agent log
    _debug_log(
        "H4",
        "webapp.py:/api/overrides/bulk",
        "Bulk overrides saved",
        {
            "project": project,
            "scope": scope,
            "incoming_rows": len(rows),
            "saved_count": saved_count,
            "failed_count": len(failed),
            "sample_item_keys": [
                (r.get("item_key") or _override_item_key((r.get("article") or ""), r.get("inventory_row")))
                for r in rows[:8]
            ],
        },
    )
    # endregion
    return jsonify({"ok": True, "saved_count": saved_count, "failed": failed})


@app.get("/api/rules")
def list_rules():
    project = request.args.get("project")
    rule_type = request.args.get("type")
    include_disabled = (request.args.get("include_disabled") or "").lower() in {"1", "true", "yes"}
    rules = store.list_rules(
        project_code=project,
        rule_type=rule_type,
        include_disabled=include_disabled,
    )
    return jsonify({"ok": True, "rules": rules})


@app.post("/api/rules")
def upsert_rule():
    payload = request.get_json(silent=True) or {}
    project = (payload.get("project_code") or "chirana").strip().lower()
    rule_type = (payload.get("rule_type") or "").strip().lower()
    scope = (payload.get("scope") or "project").strip().lower()
    rule_key = (payload.get("rule_key") or "").strip()
    rule_value = payload.get("rule_value") or {}
    reason = payload.get("reason") or ""
    priority = int(payload.get("priority") or 100)
    enabled = bool(payload.get("enabled", True))

    if rule_type not in {"alias", "category", "mapping"}:
        return jsonify({"error": "rule_type must be alias/category/mapping"}), 400
    if scope not in {"project", "system"}:
        return jsonify({"error": "scope must be project/system"}), 400
    if not rule_key:
        return jsonify({"error": "rule_key is required"}), 400
    if not isinstance(rule_value, dict):
        return jsonify({"error": "rule_value must be object"}), 400

    scope_value = "global" if scope == "system" else project
    rule_id = store.upsert_rule(
        rule_type=rule_type,
        scope_type=scope,
        scope_value=scope_value,
        rule_key=rule_key,
        rule_value=rule_value,
        reason=reason,
        priority=priority,
        enabled=enabled,
    )
    return jsonify({"ok": True, "rule_id": rule_id})


@app.delete("/api/rules/<int:rule_id>")
def delete_rule(rule_id: int):
    deleted = store.delete_rule(rule_id)
    return jsonify({"ok": True, "deleted": deleted})


@app.post("/api/rules/snapshot")
def create_rules_snapshot():
    payload = request.get_json(silent=True) or {}
    project = (payload.get("project_code") or "chirana").strip().lower()
    label = (payload.get("label") or f"snapshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}").strip()
    snapshot_id = store.create_rules_snapshot(project, label)
    return jsonify({"ok": True, "snapshot_id": snapshot_id, "label": label})


@app.get("/api/rules/snapshots")
def list_rules_snapshots():
    project = (request.args.get("project") or "chirana").strip().lower()
    snapshots = store.list_rules_snapshots(project)
    return jsonify({"ok": True, "snapshots": snapshots})


@app.get("/api/rules/export")
def export_rules():
    project = (request.args.get("project") or "chirana").strip().lower()
    rules = store.list_rules(project_code=project, include_disabled=True)
    return jsonify({"ok": True, "project_code": project, "rules": rules})


@app.post("/api/rules/import")
def import_rules():
    payload = request.get_json(silent=True) or {}
    project = (payload.get("project_code") or "chirana").strip().lower()
    scope_default = (payload.get("scope_default") or "project").strip().lower()
    rules = payload.get("rules") or []
    if not isinstance(rules, list):
        return jsonify({"error": "rules must be array"}), 400
    imported = 0
    failed: list[dict] = []
    for idx, item in enumerate(rules):
        try:
            rule_type = (item.get("rule_type") or "").strip().lower()
            scope = (item.get("scope_type") or item.get("scope") or scope_default).strip().lower()
            rule_key = (item.get("rule_key") or "").strip()
            rule_value = item.get("rule_value") or {}
            if rule_type not in {"alias", "category", "mapping"}:
                raise ValueError("invalid rule_type")
            if scope not in {"project", "system"}:
                raise ValueError("invalid scope")
            if not rule_key:
                raise ValueError("missing rule_key")
            if not isinstance(rule_value, dict):
                raise ValueError("rule_value must be object")
            scope_value = "global" if scope == "system" else project
            store.upsert_rule(
                rule_type=rule_type,
                scope_type=scope,
                scope_value=scope_value,
                rule_key=rule_key,
                rule_value=rule_value,
                reason=item.get("reason") or "",
                priority=int(item.get("priority") or 100),
                enabled=bool(item.get("enabled", True)),
            )
            imported += 1
        except Exception as exc:
            failed.append({"index": idx, "error": str(exc)})
    return jsonify({"ok": True, "imported": imported, "failed": failed})


@app.post("/api/training/files/preview")
def preview_training_files():
    _ensure_dirs()
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "At least one file is required"}), 400

    session_dir = UPLOAD_DIR / f"training_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"
    session_dir.mkdir(parents=True, exist_ok=True)
    allowed_ext = {".xlsx", ".xlsm", ".xls", ".csv"}

    payload: list[dict[str, object]] = []
    for f in files:
        name = (f.filename or "").strip()
        if not name:
            continue
        ext = Path(name).suffix.lower()
        if ext not in allowed_ext:
            payload.append(
                {
                    "filename": name,
                    "accepted": False,
                    "error": f"Unsupported extension: {ext}. Allowed: {', '.join(sorted(allowed_ext))}",
                }
            )
            continue
        saved_path = Path(_save_uploaded_file(f, session_dir))
        kind = _detect_training_file_kind(saved_path, name)
        payload.append(
            {
                "filename": name,
                "accepted": True,
                "detected_kind": kind,
                "saved_path": str(saved_path),
            }
        )
    return jsonify(
        {
            "ok": True,
            "files": payload,
            "supported_kinds": ["spp", "inventory", "writeoff", "rules", "nomenclature"],
            "help": {
                "spp": "SPP / vykaz praci za mesic",
                "inventory": "Inventura NF-30 nebo soupis/fakturace s polozkami",
                "writeoff": "NF-45 skutecne odpisy (volitelne)",
                "rules": "Excel s pravidly/aliasy",
                "nomenclature": "Referencni seznam materialu",
            },
        }
    )


@app.post("/api/analyze")
def analyze():
    _ensure_dirs()
    form = request.form
    files = request.files

    project_code = (form.get("project_code") or "chirana").strip().lower()
    project_name = (form.get("project_name") or project_code.title()).strip()
    period_month = (form.get("period_month") or datetime.now().strftime("%Y-%m")).strip()
    project_prompt = form.get("project_prompt") or ""
    no_ai = _bool(form.get("no_ai"))
    if no_ai:
        return jsonify({"error": "AI mode is required. Vypnete 'Rezim bez AI' a spustte znovu."}), 400
    auto_map = not _bool(form.get("disable_auto_map"))

    spp = files.get("spp")
    inventory = files.get("inventory")
    nf45 = files.get("nf45")
    rules_file = files.get("rules")
    if spp is None or inventory is None:
        return jsonify({"error": "spp and inventory files are required"}), 400
    slot_error = _validate_upload_slots(spp, inventory, nf45, rules_file)
    if slot_error:
        return jsonify({"error": slot_error}), 400

    session_dir = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    session_dir.mkdir(parents=True, exist_ok=True)

    spp_path = _save_uploaded_file(spp, session_dir)
    inventory_path = _save_uploaded_file(inventory, session_dir)
    nf45_path = _save_uploaded_file(nf45, session_dir) if nf45 else None
    rules_path = _save_uploaded_file(rules_file, session_dir) if rules_file else _find_default_rules_file()
    nomenclature_path = _find_default_nomenclature_file()

    try:
        spp_mapping_override = json.loads(form.get("spp_mapping") or "null")
    except json.JSONDecodeError:
        return jsonify({"error": "Invalid spp_mapping JSON"}), 400
    try:
        inv_mapping_override = json.loads(form.get("inventory_mapping") or "null")
    except json.JSONDecodeError:
        return jsonify({"error": "Invalid inventory_mapping JSON"}), 400

    store.ensure_project(project_code, project_name, project_prompt)
    if spp_mapping_override is None:
        spp_mapping_override = store.get_mapping(project_code, "spp")
    if inv_mapping_override is None:
        inv_mapping_override = store.get_mapping(project_code, "inventory")

    try:
        result = _run_pipeline(
            project_code=project_code,
            spp_path=spp_path,
            inventory_path=inventory_path,
            period_month=period_month,
            no_ai=no_ai,
            auto_map=auto_map,
            spp_mapping_override=spp_mapping_override,
            inv_mapping_override=inv_mapping_override,
            project_prompt=project_prompt,
            rules_path=rules_path,
            nomenclature_path=nomenclature_path,
            nf45_path=nf45_path,
            generate_excel=False,
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    if result["mappings"]["spp"]:
        store.save_mapping(project_code, "spp", result["mappings"]["spp"])
    if result["mappings"]["inventory"]:
        store.save_mapping(project_code, "inventory", result["mappings"]["inventory"])
    _write_draft_artifacts(spp_path, result)

    draft_id = uuid.uuid4().hex
    store.create_analysis_draft(
        draft_id=draft_id,
        project_code=project_code,
        project_name=project_name,
        period_month=period_month,
        spp_path=spp_path,
        inventory_path=inventory_path,
        nf45_path=nf45_path,
        rules_path=rules_path,
        nomenclature_path=nomenclature_path,
        project_prompt=project_prompt,
        spp_mapping=result["mappings"]["spp"],
        inventory_mapping=result["mappings"]["inventory"],
    )

    return jsonify(
        {
            "ok": True,
            "draft_id": draft_id,
            "result": _strip_export_artifacts(result),
        }
    )


@app.post("/api/review/recalculate")
def review_recalculate():
    payload = request.get_json(silent=True) or {}
    draft_id = (payload.get("draft_id") or "").strip()
    if not draft_id:
        return jsonify({"error": "draft_id is required"}), 400

    draft = store.get_analysis_draft(draft_id)
    if not draft:
        return jsonify({"error": "Draft not found"}), 404

    try:
        result = _run_pipeline(
            project_code=draft["project_code"],
            spp_path=draft["spp_path"],
            inventory_path=draft["inventory_path"],
            period_month=draft["period_month"],
            no_ai=False,
            auto_map=True,
            spp_mapping_override=draft.get("spp_mapping"),
            inv_mapping_override=draft.get("inventory_mapping"),
            project_prompt=draft.get("project_prompt") or "",
            rules_path=draft.get("rules_path"),
            nomenclature_path=draft.get("nomenclature_path"),
            nf45_path=draft.get("nf45_path"),
            generate_excel=False,
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    _write_draft_artifacts(draft["spp_path"], result)
    return jsonify({"ok": True, "draft_id": draft_id, "result": _strip_export_artifacts(result)})


@app.post("/api/review/finalize")
def review_finalize():
    payload = request.get_json(silent=True) or {}
    draft_id = (payload.get("draft_id") or "").strip()
    if not draft_id:
        return jsonify({"error": "draft_id is required"}), 400

    draft = store.get_analysis_draft(draft_id)
    if not draft:
        return jsonify({"error": "Draft not found"}), 404

    try:
        result = _finalize_from_cached_artifacts(draft)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    session_id = store.save_session(
        project_code=draft["project_code"],
        period_month=draft["period_month"],
        spp_path=draft["spp_path"],
        inventory_path=draft["inventory_path"],
        nf45_path=draft.get("nf45_path"),
        output_path=result["output_path"],
        stats=result,
    )
    return jsonify({"ok": True, "session_id": session_id, "result": result})


@app.post("/api/detect-mapping")
def detect_mapping():
    files = request.files
    spp = files.get("spp")
    inventory = files.get("inventory")
    if spp is None and inventory is None:
        return jsonify({"error": "spp or inventory file is required"}), 400

    _ensure_dirs()
    session_dir = UPLOAD_DIR / f"mapping_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"
    session_dir.mkdir(parents=True, exist_ok=True)

    api_key = (
        settings.OPENAI_API_KEY
        or os.environ.get("OPENAI_API_KEY")
        or settings.ANTHROPIC_API_KEY
        or os.environ.get("ANTHROPIC_API_KEY")
        or ""
    ).strip()
    client = ClaudeClient(api_key=api_key) if api_key else None

    result = {}
    if spp:
        spp_path = Path(_save_uploaded_file(spp, session_dir))
        mapping = _default_mapping("spp")
        if client:
            try:
                mapping = auto_detect_mapping(spp_path, client, "spp")
            except Exception:
                pass
        result["spp_mapping"] = mapping_to_dict(mapping)
    if inventory:
        inv_path = Path(_save_uploaded_file(inventory, session_dir))
        mapping = _default_mapping("inventory")
        if client:
            try:
                mapping = auto_detect_mapping(inv_path, client, "inventory")
            except Exception:
                pass
        result["inventory_mapping"] = mapping_to_dict(mapping)

    return jsonify({"ok": True, **result})


@app.get("/output/<path:filename>")
def get_output(filename: str):
    return send_from_directory(str(OUTPUT_DIR), filename, as_attachment=True)


if __name__ == "__main__":
    _ensure_dirs()
    app.run(host="127.0.0.1", port=8000, debug=False, threaded=True)
