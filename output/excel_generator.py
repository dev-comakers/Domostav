"""Generate output Excel file with AI analysis columns."""

from __future__ import annotations

import json
import shutil
from pathlib import Path
import time
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import AnomalyStatus, MatchMethod, WriteoffRecommendation


_DEBUG_LOG_PATH = "/Users/dmytriivezerian/Desktop/Domostav x Fajnwork/.cursor/debug-f07731.log"


def _debug_log(hypothesis_id: str, location: str, message: str, data: dict[str, Any]) -> None:
    payload = {
        "sessionId": "f07731",
        "runId": "initial",
        "hypothesisId": hypothesis_id,
        "location": location,
        "message": message,
        "data": data,
        "timestamp": int(time.time() * 1000),
    }
    try:
        with open(_DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        pass


FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_WARNING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_REVIEW = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_NORMAL = Font(size=9)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

AI_COLUMNS = [
    ("AI: Ожид. списание", 15),
    ("AI: Ед.", 9),
    ("AI: Привязка к SPP", 40),
    ("AI: Источник SPP", 22),
    ("AI: Причина", 35),
    ("AI: Статус", 12),
    ("AI: Метод", 12),
    ("AI: Откл. %", 12),
]


def _pick_td_sheet_name(wb: openpyxl.Workbook) -> str:
    preferred_tokens = ("asr", "kanal", "vodo", "vytap", "chlazen", "zavlah")
    excluded_names = {"ai summary", "spp coverage"}
    candidates: list[tuple[int, str]] = []
    for name in wb.sheetnames:
        low = name.strip().lower()
        if low in excluded_names:
            continue
        ws = wb[name]
        score = 0
        if any(tok in low for tok in preferred_tokens):
            score += 100
        if "rekap" in low or "summary" in low:
            score -= 60
        score += min(ws.max_row, 400) // 10
        score += min(ws.max_column, 80)
        candidates.append((score, name))
    if not candidates:
        return wb.active.title
    candidates.sort(reverse=True)
    return candidates[0][1]


def _ensure_td_sheet_from_source(
    wb: openpyxl.Workbook,
    requested_sheet_name: str | None,
) -> tuple[openpyxl.worksheet.worksheet.Worksheet, str]:
    source_name = requested_sheet_name if requested_sheet_name in wb.sheetnames else _pick_td_sheet_name(wb)
    source_ws = wb[source_name] if source_name in wb.sheetnames else wb.active
    if "TDSheet" in wb.sheetnames:
        del wb["TDSheet"]
    td_ws = wb.copy_worksheet(source_ws)
    td_ws.title = "TDSheet"
    return td_ws, source_ws.title


def _human_method_label(method: MatchMethod) -> str:
    if method == MatchMethod.MANUAL:
        return "Manual override"
    if method == MatchMethod.ARTICLE:
        return "Article match"
    if method == MatchMethod.REGEX:
        return "Regex match"
    if method == MatchMethod.UNMATCHED:
        return "No match"
    return "AI detection"


def _detect_unit_column(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> int | None:
    unit_keywords = ("jed", "unit", "ед", "mj", "měr")
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue
        text = str(value).strip().lower()
        if any(k in text for k in unit_keywords):
            return col
    # Chirana default
    if ws.max_column >= 20:
        return 20
    return None


def generate_output(
    source_path: str | Path,
    output_path: str | Path,
    recommendations: list[WriteoffRecommendation],
    data_start_row: int = 12,
    sheet_name: str | None = None,
    summary: dict | None = None,
    spp_coverage: list[dict] | None = None,
) -> Path:
    """Generate output by cloning inventory file and appending AI columns."""
    started_at = time.perf_counter()
    source_path = Path(source_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # region agent log
    _debug_log(
        "H3",
        "output/excel_generator.py:88",
        "Excel generation started",
        {
            "source_path": str(source_path),
            "output_path": str(output_path),
            "recommendation_count": len(recommendations),
            "has_summary": bool(summary),
            "spp_coverage_count": len(spp_coverage or []),
        },
    )
    # endregion

    shutil.copy2(source_path, output_path)

    wb = openpyxl.load_workbook(str(output_path))
    ws, source_sheet_name = _ensure_td_sheet_from_source(wb, sheet_name)
    # region agent log
    _debug_log(
        "H2",
        "output/excel_generator.py:ws-selection",
        "Excel target sheet selected",
        {
            "requested_sheet_name": sheet_name,
            "auto_selected_sheet_name": source_sheet_name,
            "selected_sheet_name": ws.title,
            "workbook_sheet_names": list(wb.sheetnames),
            "data_start_row": data_start_row,
            "ws_max_row": ws.max_row,
            "ws_max_col": ws.max_column,
        },
    )
    # endregion

    ai_start_col = ws.max_column + 2
    header_row = data_start_row - 1
    unit_col = _detect_unit_column(ws, header_row)

    for i, (col_name, width) in enumerate(AI_COLUMNS):
        col_idx = ai_start_col + i
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for rec in recommendations:
        row = rec.inventory_row
        col = ai_start_col

        # AI: Ожид. списание
        cell = ws.cell(row=row, column=col, value=rec.expected_writeoff)
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.number_format = "#,##0.00"
        col += 1

        # AI: Ед.
        unit_value = ""
        if unit_col is not None:
            raw_unit = ws.cell(row=row, column=unit_col).value
            unit_value = str(raw_unit).strip() if raw_unit is not None else ""
        unit_cell = ws.cell(row=row, column=col, value=unit_value)
        unit_cell.font = FONT_NORMAL
        unit_cell.border = THIN_BORDER
        unit_cell.alignment = Alignment(horizontal="center")
        col += 1

        # AI: Привязка к SPP
        cell = ws.cell(
            row=row, column=col,
            value=rec.spp_reference[:200] if rec.spp_reference else "",
        )
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)
        col += 1

        # AI: Источник SPP — compact form like "ZTI #6, UT #12"
        source_label = _extract_spp_source(rec.spp_reference)
        cell = ws.cell(row=row, column=col, value=source_label)
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)
        col += 1

        # AI: Причина
        cell = ws.cell(row=row, column=col, value=rec.reason)
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)
        col += 1

        # AI: Статус
        status_cell = ws.cell(row=row, column=col, value=rec.status.value)
        status_cell.font = FONT_NORMAL
        status_cell.border = THIN_BORDER
        status_cell.alignment = Alignment(horizontal="center")
        if rec.status == AnomalyStatus.OK:
            status_cell.fill = FILL_OK
        elif rec.status == AnomalyStatus.WARNING:
            status_cell.fill = FILL_WARNING
        else:
            status_cell.fill = FILL_RED
        col += 1

        # AI: Метод
        cell = ws.cell(row=row, column=col, value=_human_method_label(rec.match_method))
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        col += 1

        # AI: Откл. %
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.number_format = "0.0%"
        if rec.deviation_percent is not None:
            cell.value = rec.deviation_percent / 100

    _add_summary_sheet(wb, recommendations, summary=summary)
    _add_spp_coverage_sheet(wb, spp_coverage or [])
    sheet_names = list(wb.sheetnames)
    wb.save(str(output_path))
    wb.close()

    # region agent log
    _debug_log(
        "H3",
        "output/excel_generator.py:202",
        "Excel generation finished",
        {
            "output_path": str(output_path),
            "elapsed_ms": round((time.perf_counter() - started_at) * 1000, 1),
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
        },
    )
    # endregion
    return output_path


def _add_summary_sheet(
    wb: openpyxl.Workbook,
    recommendations: list[WriteoffRecommendation],
    summary: dict | None = None,
) -> None:
    if "AI Summary" in wb.sheetnames:
        del wb["AI Summary"]
    ws = wb.create_sheet("AI Summary")

    summary = summary or {}
    total = int(summary.get("total_items", len(recommendations)))
    ok = int(summary.get("ok", sum(1 for r in recommendations if r.status == AnomalyStatus.OK)))
    warning = int(summary.get("warning", sum(1 for r in recommendations if r.status == AnomalyStatus.WARNING)))
    red = int(summary.get("red_flag", sum(1 for r in recommendations if r.status == AnomalyStatus.RED_FLAG)))
    review = int(summary.get("review", sum(1 for r in recommendations if r.expected_writeoff is None)))
    money_totals = summary.get("money_totals", {})
    top_anomalies = summary.get("top_anomalies", [])

    ws.cell(row=1, column=1, value="Итоги списания материалов — версия для руководителя").font = Font(
        bold=True, size=14
    )

    # Block 1: count of positions
    ws.cell(row=3, column=1, value="1) По количеству позиций").font = Font(bold=True, size=11)
    ws.cell(row=4, column=1, value="Всего позиций:")
    ws.cell(row=4, column=2, value=total)
    ws.cell(row=5, column=1, value="OK (норма):")
    ws.cell(row=5, column=2, value=ok)
    ws.cell(row=5, column=3).fill = FILL_OK
    ws.cell(row=6, column=1, value="WARNING (нужно проверить):")
    ws.cell(row=6, column=2, value=warning)
    ws.cell(row=6, column=3).fill = FILL_WARNING
    ws.cell(row=7, column=1, value="REVIEW (не привязано к SPP):")
    ws.cell(row=7, column=2, value=review)
    ws.cell(row=7, column=3).fill = FILL_REVIEW
    ws.cell(row=8, column=1, value="RED FLAG (критично):")
    ws.cell(row=8, column=2, value=red)
    ws.cell(row=8, column=3).fill = FILL_RED

    # Block 3: money
    ws.cell(row=10, column=1, value="3) Денежный итог (если есть цена)").font = Font(bold=True, size=11)
    ws.cell(row=10, column=4, value="Формула: количество * цена по каждой позиции").font = Font(italic=True, size=9)
    ws.cell(row=11, column=1, value="Ожидаемая сумма списания:")
    ws.cell(row=11, column=2, value=float(money_totals.get("expected_cost", 0.0)))
    ws.cell(row=12, column=1, value="Фактическая сумма (по отклонению):")
    ws.cell(row=12, column=2, value=float(money_totals.get("actual_cost", 0.0)))
    ws.cell(row=13, column=1, value="Разница факт - ожидание:")
    ws.cell(row=13, column=2, value=float(money_totals.get("delta_cost", 0.0)))
    ws.cell(row=14, column=1, value="Ед. измерения денег: в валюте цен из файла (обычно CZK).").font = Font(italic=True, size=9)

    # Block 5: top anomalies one-line explanation
    top_row = 17
    ws.cell(row=top_row, column=1, value="5) Топ аномалий (1 строка для заказчика)").font = Font(bold=True, size=11)
    headers = ["Строка", "Материал", "Ед.", "Ожидалось", "Факт", "Влияние, CZK", "Статус", "Пояснение (1 строка)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=top_row + 1, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER

    def _money_impact_from_rec(rec: WriteoffRecommendation) -> float:
        if rec.actual_deviation is None:
            return 0.0
        # Fallback approximation when per-row price is unavailable in summary payload.
        return abs(float(rec.actual_deviation))

    if not top_anomalies:
        red_flags = [r for r in recommendations if r.status == AnomalyStatus.RED_FLAG]
        red_flags.sort(
            key=lambda r: _money_impact_from_rec(r),
            reverse=True,
        )
        for idx, rec in enumerate(red_flags[:10], top_row + 2):
            ws.cell(row=idx, column=1, value=rec.inventory_row)
            ws.cell(row=idx, column=2, value=rec.inventory_name[:60])
            ws.cell(row=idx, column=3, value="")
            ws.cell(row=idx, column=4, value=rec.expected_writeoff)
            ws.cell(row=idx, column=5, value=rec.actual_deviation)
            ws.cell(row=idx, column=6, value=_money_impact_from_rec(rec))
            ws.cell(row=idx, column=7, value=rec.status.value)
            ws.cell(row=idx, column=8, value=_human_reason_fallback(rec))
    else:
        top_anomalies = sorted(
            top_anomalies,
            key=lambda item: abs(float(item.get("money_impact") or 0.0)),
            reverse=True,
        )
        for idx, item in enumerate(top_anomalies[:10], top_row + 2):
            ws.cell(row=idx, column=1, value=item.get("row"))
            ws.cell(row=idx, column=2, value=str(item.get("name", ""))[:60])
            ws.cell(row=idx, column=3, value=item.get("unit"))
            ws.cell(row=idx, column=4, value=item.get("expected_writeoff"))
            ws.cell(row=idx, column=5, value=item.get("actual_deviation"))
            ws.cell(row=idx, column=6, value=item.get("money_impact"))
            ws.cell(row=idx, column=7, value=item.get("status"))
            ws.cell(row=idx, column=8, value=item.get("one_line_explanation") or str(item.get("reason", ""))[:200])

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 78

    ws.cell(
        row=2,
        column=1,
        value="Важно: m, ks и kg считаются отдельно; общий итог по количеству между единицами не смешивается.",
    ).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)


def _extract_spp_source(spp_reference: str) -> str:
    """Turn '[ZTI] Row 6: ...' into compact 'ZTI #6'."""
    import re
    if not spp_reference:
        return ""
    parts = re.findall(r"\[([^\]]+)\]\s*Row\s*(\d+)", spp_reference)
    if not parts:
        return spp_reference[:40]
    return ", ".join(f"{sheet} #{row}" for sheet, row in parts)


def _add_spp_coverage_sheet(
    wb: openpyxl.Workbook,
    coverage: list[dict],
) -> None:
    """Write 'SPP Coverage' sheet: each row = one SPP item for the month."""
    if "SPP Coverage" in wb.sheetnames:
        del wb["SPP Coverage"]
    ws = wb.create_sheet("SPP Coverage")

    FILL_COVERED = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FILL_NOT = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.cell(row=1, column=1, value="Pokryti SPP — kontrola kazdé polozky SPP za mesic").font = Font(
        bold=True, size=14,
    )
    ws.cell(row=2, column=1, value=(
        "Kazdy radek = prace z SPP za zvoleny mesic. "
        "Sloupec 'Pokryto?' ukazuje, zda sklad pokryl tuto praci."
    )).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)

    headers = [
        ("List SPP", 12),
        ("Radek", 8),
        ("Prace (nazev)", 52),
        ("Jednotka", 9),
        ("Plan za mesic (mnozstvi)", 18),
        ("Plan za mesic (CZK)", 18),
        ("Pokryto?", 12),
        ("Radky skladu", 14),
        ("Materialy ze skladu", 48),
        ("Fakt ze skladu", 16),
        ("Rozdil (plan-fakt)", 18),
        ("Komentar", 56),
    ]
    for i, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w

    covered_count = 0
    not_covered_count = 0

    for idx, item in enumerate(coverage, 5):
        ws.cell(row=idx, column=1, value=item["spp_sheet"]).border = THIN_BORDER
        ws.cell(row=idx, column=2, value=item["spp_row"]).border = THIN_BORDER
        ws.cell(row=idx, column=3, value=item["spp_name"][:100]).border = THIN_BORDER
        ws.cell(row=idx, column=4, value=item.get("spp_unit") or "").border = THIN_BORDER

        qty_cell = ws.cell(row=idx, column=5, value=item["spp_qty_month"])
        qty_cell.number_format = "#,##0.00"
        qty_cell.border = THIN_BORDER

        tm_cell = ws.cell(row=idx, column=6, value=item.get("spp_total_month"))
        tm_cell.number_format = "#,##0.00"
        tm_cell.border = THIN_BORDER

        covered_val = "Ano" if item["covered"] else "Ne"
        cov_cell = ws.cell(row=idx, column=7, value=covered_val)
        cov_cell.border = THIN_BORDER
        cov_cell.alignment = Alignment(horizontal="center")
        if item["covered"]:
            cov_cell.fill = FILL_COVERED
            covered_count += 1
        else:
            cov_cell.fill = FILL_NOT
            not_covered_count += 1

        inv_rows_str = ", ".join(str(r) for r in item["inventory_rows"][:10])
        ws.cell(row=idx, column=8, value=inv_rows_str).border = THIN_BORDER

        inv_names_str = "; ".join(item["inventory_names"][:5])
        ws.cell(row=idx, column=9, value=inv_names_str[:200]).border = THIN_BORDER

        dev_cell = ws.cell(row=idx, column=10, value=item["inventory_total_deviation"])
        dev_cell.number_format = "#,##0.00"
        dev_cell.border = THIN_BORDER

        delta_cell = ws.cell(row=idx, column=11, value=item["delta"])
        delta_cell.number_format = "#,##0.00"
        delta_cell.border = THIN_BORDER

        ws.cell(row=idx, column=12, value=item["comment"]).border = THIN_BORDER

    total_row = len(coverage) + 6
    ws.cell(row=total_row, column=1, value="CELKEM").font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=7, value=f"Ano: {covered_count}, Ne: {not_covered_count}").font = Font(bold=True)
    total_items = max(len(coverage), 1)
    ws.cell(
        row=total_row + 1, column=1,
        value=f"Pokryti: {covered_count}/{total_items} ({round(covered_count/total_items*100, 1)}%)",
    ).font = Font(bold=True, size=11)

    # Per-sheet summary (e.g., ZTI vs UT), to avoid confusion about totals.
    by_sheet: dict[str, dict[str, int]] = {}
    for item in coverage:
        sheet = str(item.get("spp_sheet") or "UNKNOWN")
        block = by_sheet.setdefault(sheet, {"total": 0, "covered": 0, "not_covered": 0})
        block["total"] += 1
        if item.get("covered"):
            block["covered"] += 1
        else:
            block["not_covered"] += 1

    start = total_row + 3
    ws.cell(row=start, column=1, value="Pokryti po listech SPP").font = Font(bold=True, size=11)
    ws.cell(row=start + 1, column=1, value="List").font = FONT_HEADER
    ws.cell(row=start + 1, column=2, value="Pokryto").font = FONT_HEADER
    ws.cell(row=start + 1, column=3, value="Nepokryto").font = FONT_HEADER
    ws.cell(row=start + 1, column=4, value="Celkem").font = FONT_HEADER
    for c in range(1, 5):
        ws.cell(row=start + 1, column=c).fill = FILL_HEADER
        ws.cell(row=start + 1, column=c).border = THIN_BORDER

    line = start + 2
    for sheet in sorted(by_sheet.keys()):
        stats = by_sheet[sheet]
        ws.cell(row=line, column=1, value=sheet).border = THIN_BORDER
        ws.cell(row=line, column=2, value=stats["covered"]).border = THIN_BORDER
        ws.cell(row=line, column=3, value=stats["not_covered"]).border = THIN_BORDER
        ws.cell(row=line, column=4, value=stats["total"]).border = THIN_BORDER
        line += 1


def _human_reason_fallback(rec: WriteoffRecommendation) -> str:
    expected = f"{rec.expected_writeoff:.1f}" if rec.expected_writeoff is not None else "нет оценки"
    actual = f"{rec.actual_deviation:.1f}" if rec.actual_deviation is not None else "нет факта"
    pct = f"{rec.deviation_percent:.1f}%" if rec.deviation_percent is not None else "нет %"
    return (
        f"Материал: {rec.inventory_name[:45]}. Ожидалось: {expected}. "
        f"Факт: {actual}. Отклонение: {pct}. Статус: {rec.status.value}."
    )
