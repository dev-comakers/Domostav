from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .company_rules import is_domostav_company


HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
TOTAL_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)

MONTH_LABELS_CS = {
    "01": "Leden", "02": "\u00danor", "03": "B\u0159ezen", "04": "Duben",
    "05": "Kv\u011bten", "06": "\u010cerven", "07": "\u010cervenec", "08": "Srpen",
    "09": "Z\u00e1\u0159\u00ed", "10": "\u0158\u00edjen", "11": "Listopad", "12": "Prosinec",
}


# ---------------------------------------------------------------------------
# Column schemas (matching Svetlana's Vika_Sasha reference template)
# ---------------------------------------------------------------------------

# Main format — for every company EXCEPT Domostav (FT, KOPA, BAU, NR, BSS, OKI, Au...)
# Grouped by Projekt. Per-row Celkem formula. Subtotal row per project.
MAIN_HEADERS = [
    "P\u0159\u00edjmen\u00ed",
    "Projekt",
    "Koordin\u00e1tor",
    "Odvody plat\u00edme",
    "Odvody strh\u00e1v\u00e1me",
    "Bonus 1",
    "Bonus 2",
    "Na \u00fa\u010det (pouze LEFT OF)",
    "Celkem",
    "M\u011bs\u00ed\u010dn\u00ed mzda",
    "Firma",
]
MAIN_WIDTHS = [28, 26, 18, 16, 18, 12, 12, 22, 14, 16, 22]

# DM format — Domostav companies. Numbered column B with =B(n-1)+1 formula.
# No Koordinátor column. Per-row Celkem mirrors the main sheet formula.
DM_HEADERS = [
    "",
    "\u010c\u00edslo",
    "P\u0159\u00edjmen\u00ed",
    "Projekt",
    "Odvody plat\u00edme",
    "Odvody strh\u00e1v\u00e1me",
    "Bonus 1",
    "Bonus 2",
    "Na \u00fa\u010det",
    "Celkem",
    "M\u011bs\u00ed\u010dn\u00ed mzda",
    "Firma",
]
DM_WIDTHS = [3, 8, 30, 24, 16, 18, 12, 12, 16, 14, 16, 22]


# ---------------------------------------------------------------------------
# Row-level helpers
# ---------------------------------------------------------------------------


def is_domostav_row(row: dict) -> bool:
    return is_domostav_company(row.get("company_name"), row.get("company_code"))


def company_label(row: dict) -> str:
    name = (row.get("company_name") or "").strip()
    if name:
        return name
    return (row.get("company_code") or "").strip()


def project_label(row: dict) -> str:
    return (row.get("project_name") or "").strip() or "Bez projektu"


def month_year_label(period: str) -> str:
    if not period or "/" not in period:
        now = datetime.now()
        return f"{MONTH_LABELS_CS.get(f'{now.month:02d}', '')} {now.year}".strip()
    month, year = period.split("/", 1)
    return f"{MONTH_LABELS_CS.get(month, month)} {year}".strip()


def sheet_title_for_period(period: str) -> str:
    return (period.replace("/", ".") if period else "export")[:31]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def detect_export_variant(preview_rows: list[dict]) -> str:
    """Return 'dm' if the import is a Domostav batch, 'main' otherwise.

    Mixed batches default to 'main' (but Tati\u00e1na's real workflow keeps them separate).
    """

    if not preview_rows:
        return "main"
    domostav_count = sum(1 for r in preview_rows if is_domostav_row(r))
    if domostav_count == 0:
        return "main"
    if domostav_count == len(preview_rows):
        return "dm"
    # mixed batch fallback: treat as main layout (which has the Firma column)
    return "main"


def build_export(preview_rows: list[dict], output_path: str) -> str:
    """Generate one XLSX matching the Vika_Sasha reference layout.

    Layout is picked automatically based on the contents of the import:
    - if every row is Domostav -> DM layout (numbered column, no coordinator)
    - otherwise               -> main layout (all companies grouped by project)
    """

    if not preview_rows:
        raise ValueError("Nothing to export")

    period = preview_rows[0].get("period") or ""
    variant = detect_export_variant(preview_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title_for_period(period) if variant == "main" else "DM"

    if variant == "dm":
        _build_dm_sheet(ws, preview_rows, period)
    else:
        _build_main_sheet(ws, preview_rows, period)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Main sheet (all non-Domostav companies)
# ---------------------------------------------------------------------------


def _build_main_sheet(ws, rows: list[dict], period: str) -> None:
    ws["A1"] = month_year_label(period)
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(MAIN_HEADERS))

    for idx, header in enumerate(MAIN_HEADERS, start=1):
        cell = ws.cell(3, idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for idx, width in enumerate(MAIN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "A4"

    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[project_label(row)].append(row)

    current = 4
    for project in _sort_projects(grouped.keys()):
        group_start = current
        for item in sorted(grouped[project], key=lambda r: r.get("display_name") or ""):
            ws.cell(current, 1).value = (item.get("display_name") or "").strip()
            ws.cell(current, 2).value = project
            ws.cell(current, 3).value = item.get("coordinator_name") or None
            ws.cell(current, 4).value = float(item.get("odvody_platime") or 0)
            ws.cell(current, 5).value = float(item.get("odvody_strhavame") or 0)
            ws.cell(current, 6).value = None  # Bonus 1 — ruční
            ws.cell(current, 7).value = None  # Bonus 2 — ruční
            ws.cell(current, 8).value = None  # Na účet — ruční
            ws.cell(current, 9).value = f"=E{current}+F{current}+G{current}+H{current}"
            mesicni = float(item.get("mesicni_mzda") or 0)
            ws.cell(current, 10).value = mesicni if mesicni > 0 else None
            ws.cell(current, 11).value = company_label(item) or None

            for col in ("D", "E", "F", "G", "H", "I", "J"):
                ws[f"{col}{current}"].number_format = "#,##0.00"
            current += 1

        group_end = current - 1
        _write_subtotal(
            ws,
            current,
            project,
            group_start,
            group_end,
            name_col=1,
            sum_cols=("D", "E", "F", "G", "H", "I"),
            total_col_count=len(MAIN_HEADERS),
        )
        current += 2


# ---------------------------------------------------------------------------
# DM sheet (only Domostav)
# ---------------------------------------------------------------------------


def _build_dm_sheet(ws, rows: list[dict], period: str) -> None:
    ws["A1"] = f"DM \u2014 {month_year_label(period)}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DM_HEADERS))

    for idx, header in enumerate(DM_HEADERS, start=1):
        cell = ws.cell(3, idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for idx, width in enumerate(DM_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "A4"

    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[project_label(row)].append(row)

    current = 4
    running_index = 0
    for project in _sort_projects(grouped.keys()):
        group_start = current
        for item in sorted(grouped[project], key=lambda r: r.get("display_name") or ""):
            running_index += 1
            ws.cell(current, 1).value = None
            if running_index == 1:
                ws.cell(current, 2).value = 1
            else:
                ws.cell(current, 2).value = f"=B{current - 1}+1"
            ws.cell(current, 3).value = (item.get("display_name") or "").strip()
            ws.cell(current, 4).value = project
            ws.cell(current, 5).value = float(item.get("odvody_platime") or 0)
            ws.cell(current, 6).value = float(item.get("odvody_strhavame") or 0)
            ws.cell(current, 7).value = None  # Bonus 1
            ws.cell(current, 8).value = None  # Bonus 2
            ws.cell(current, 9).value = None  # Na účet
            ws.cell(current, 10).value = f"=F{current}+G{current}+H{current}+I{current}"
            mesicni = float(item.get("mesicni_mzda") or 0)
            ws.cell(current, 11).value = mesicni if mesicni > 0 else None
            ws.cell(current, 12).value = company_label(item) or "DOMOSTAV TZB a.s."

            for col in ("E", "F", "G", "H", "I", "J", "K"):
                ws[f"{col}{current}"].number_format = "#,##0.00"
            current += 1

        group_end = current - 1
        _write_subtotal(
            ws,
            current,
            project,
            group_start,
            group_end,
            name_col=3,
            sum_cols=("E", "F", "G", "H", "I", "J"),
            total_col_count=len(DM_HEADERS),
        )
        current += 2


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


PROJECT_ORDER_HINT = [
    "BACK OFFICE",
    "LEFT OF",
    "LEFT OFF",
    "LEFT OFFICE",
    "PALMA",
    "OK-BE",
    "HELP",
    "Instalace",
    "Stavba",
    "Elektro",
    "Domostav",
    "Domostav/ HPP",
    "Domostav / HPP",
]


def _sort_projects(projects) -> list[str]:
    def rank(p: str) -> tuple[int, str]:
        low = p.lower()
        for idx, hint in enumerate(PROJECT_ORDER_HINT):
            if low.startswith(hint.lower()):
                return (idx, low)
        return (len(PROJECT_ORDER_HINT), low)

    return sorted(projects, key=rank)


def _write_subtotal(
    ws,
    row_index: int,
    project: str,
    group_start: int,
    group_end: int,
    *,
    name_col: int,
    sum_cols: tuple[str, ...],
    total_col_count: int,
) -> None:
    label_cell = ws.cell(row_index, name_col)
    label_cell.value = f"Celkem {project}"
    label_cell.font = TOTAL_FONT
    label_cell.fill = TOTAL_FILL

    for col in sum_cols:
        cell = ws[f"{col}{row_index}"]
        cell.value = f"=SUM({col}{group_start}:{col}{group_end})"
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = "#,##0.00"

    for col_idx in range(1, total_col_count + 1):
        cell = ws[f"{get_column_letter(col_idx)}{row_index}"]
        if cell.fill.fgColor.rgb != "00D9E2F3":
            cell.fill = TOTAL_FILL
