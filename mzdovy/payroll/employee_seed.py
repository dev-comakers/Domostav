from __future__ import annotations

import re
from pathlib import Path
from typing import Any, BinaryIO

import openpyxl

from .company_rules import is_domostav_tzb

SKIP_KEYWORDS = (
    "celkem",
    "total",
    "prijmeno",
    "prijmen",
    "jm\u00e9no",
    "firma",
    "projekt",
)


_XLSX_STOP_KEYWORDS = re.compile(
    r"""
    \s+
    (?:
        NV | ZK | NAR |
        r\.n\. |
        malyi | maly | mlady | mlada | stary | stara |
        viza | neschopenka | neschopen |
        z\s+FT | z\s+DM | z\s+KOPA | z\s+OKI
    )
    \b
    """,
    re.IGNORECASE | re.VERBOSE,
)


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        text = str(value)
    else:
        text = str(value)
    text = text.strip()
    return text or None


def _money(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_skip_name(name: str) -> bool:
    lower = name.lower()
    return any(lower.startswith(keyword) for keyword in SKIP_KEYWORDS)


def clean_seed_name(raw: str) -> tuple[str, str | None]:
    if not raw:
        return "", None
    original = raw.replace("\xa0", " ").strip()
    s = original
    s = re.sub(r"/[^/]*/", " ", s)
    s = _XLSX_STOP_KEYWORDS.split(s, maxsplit=1)[0]
    s = s.split("(", 1)[0]
    s = s.split("/", 1)[0]
    s = re.split(r"\s+\d", s, maxsplit=1)[0]
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    notes = original if s != original else None
    return s, notes


def _extract_from_main_sheet(ws) -> list[dict[str, Any]]:
    """Main XLSX layout (as used by Tati\u00e1na):

    Col 0  = Jm\u00e9no
    Col 1  = Projekt
    Col 2  = Koordin\u00e1tor
    Col 3  = Odvody plat\u00edme (per-month, skip)
    Col 4  = Odvody strh\u00e1v\u00e1me (seed it as employee attribute)
    Col 5-8= Bonusy / Na \u00fa\u010det (skip, per-month)
    Col 9  = M\u011bs\u00ed\u010dn\u00ed mzda (seed it)
    Col 10 = Firma (code or name)
    """

    items: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        raw = _clean(row[0] if len(row) > 0 else None)
        if not raw or _is_skip_name(raw):
            continue
        clean_name, notes = clean_seed_name(raw)
        if not clean_name or _is_skip_name(clean_name):
            continue
        project = _clean(row[1] if len(row) > 1 else None)
        coordinator = _clean(row[2] if len(row) > 2 else None)
        odvody_strhavame = _money(row[4] if len(row) > 4 else None)
        mesicni_mzda = _money(row[9] if len(row) > 9 else None)
        company_code = _clean(row[10] if len(row) > 10 else None)
        items.append(
            {
                "full_name": clean_name,
                "project_name": project,
                "coordinator_name": coordinator,
                "company_code": company_code,
                "company_name": None,
                "notes": notes,
                "odvody_strhavame": odvody_strhavame,
                "mesicni_mzda": mesicni_mzda,
            }
        )
    return items


def _extract_from_dm_sheet(ws) -> list[dict[str, Any]]:
    """DM sheet layout:

    Col 0  = (empty)
    Col 1  = Index (number)
    Col 2  = Jm\u00e9no
    Col 3  = Projekt
    Col 4  = Odvody plat\u00edme (skip)
    Col 5  = Odvody strh\u00e1v\u00e1me (seed)
    Col 6-9= Bonusy / Na \u00fa\u010det (skip)
    Col 10 = M\u011bs\u00ed\u010dn\u00ed mzda (seed)
    Col 11 = Firma / c\u00edl
    """

    items: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        raw = _clean(row[2] if len(row) > 2 else None)
        if not raw or _is_skip_name(raw):
            continue
        clean_name, notes = clean_seed_name(raw)
        if not clean_name or _is_skip_name(clean_name):
            continue
        project = _clean(row[3] if len(row) > 3 else None)
        odvody_strhavame = _money(row[5] if len(row) > 5 else None)
        mesicni_mzda = _money(row[10] if len(row) > 10 else None)
        company_name = _clean(row[11] if len(row) > 11 else None) or "DOMOSTAV TZB a.s."
        company_code = "DM" if is_domostav_tzb(company_name, "DM") else _clean(row[11] if len(row) > 11 else None)
        items.append(
            {
                "full_name": clean_name,
                "project_name": project,
                "coordinator_name": None,
                "company_code": company_code,
                "company_name": company_name,
                "notes": notes,
                "odvody_strhavame": odvody_strhavame,
                "mesicni_mzda": mesicni_mzda,
            }
        )
    return items


def load_employees_from_xlsx(source: str | Path | BinaryIO) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    items: list[dict[str, Any]] = []

    main_sheet = wb.worksheets[0]
    items.extend(_extract_from_main_sheet(main_sheet))

    dm_sheet = None
    for ws in wb.worksheets:
        if ws.title.strip().lower() == "dm":
            dm_sheet = ws
            break
    if dm_sheet is not None:
        items.extend(_extract_from_dm_sheet(dm_sheet))

    wb.close()
    return items
